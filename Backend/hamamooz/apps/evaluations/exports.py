from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .catalog import DOMAIN_DEFINITIONS, METRIC_CATALOG
from .services import EvaluationAnalyticsService

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Vazirmatn", bold=True, color="FFFFFF")


def _safe_excel_text(value):
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value


def _style_sheet(sheet) -> None:
    sheet.sheet_view.rightToLeft = True
    sheet.freeze_panes = "A2"
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for column in sheet.columns:
        letter = column[0].column_letter
        sheet.column_dimensions[letter].width = min(
            40,
            max(12, max(len(str(cell.value or "")) for cell in column) + 2),
        )


def build_evaluation_analytics_workbook(enrollments, cohort_summary) -> BytesIO:
    workbook = Workbook()
    raw = workbook.active
    raw.title = "ثبت اطلاعات"
    metric_codes = list(METRIC_CATALOG)
    raw.append(
        ["شناسه ثبت‌نام", "نام و نام خانوادگی", "شماره دانش‌آموزی", "کلاس", "شماره ماه"]
        + [METRIC_CATALOG[code]["title"] for code in metric_codes]
        + [f"میانگین {title} (۲۰)" for title, _weight in DOMAIN_DEFINITIONS.values()]
        + ["نمره نهایی", "وضعیت تکمیل", "سطح عملکرد", "توضیحات"]
    )
    for enrollment in enrollments:
        for evaluation in EvaluationAnalyticsService._evaluations_for(enrollment):
            summary = EvaluationAnalyticsService.evaluation_summary(evaluation)
            score_map = {score.metric_code: score.value for score in evaluation.metric_scores.all()}
            raw.append(
                [
                    str(enrollment.id),
                    _safe_excel_text(enrollment.student.full_name),
                    _safe_excel_text(enrollment.student_number),
                    _safe_excel_text(enrollment.class_section.title),
                    evaluation.month_no,
                ]
                + [score_map.get(code) for code in metric_codes]
                + [item["score"] for item in summary["domain_scores"]]
                + [
                    summary["overall_score"],
                    summary["completion_status"],
                    summary["performance_level"],
                    _safe_excel_text(evaluation.note),
                ]
            )
    _style_sheet(raw)

    analysis = workbook.create_sheet("تحلیل دانش آموزان")
    analysis.append(
        [
            "شناسه ثبت‌نام",
            "نام و نام خانوادگی",
            "شماره دانش‌آموزی",
            "کلاس",
            "وضعیت تکمیل",
            "درصد تکمیل",
            "میانگین کل",
            "سطح عملکرد",
            "اولین ماه",
            "آخرین ماه",
            "میزان تغییر",
            "روند",
            "قوی‌ترین حوزه",
            "ضعیف‌ترین حوزه",
            "رتبه",
            "تعداد رتبه‌بندی‌شده",
            "پیشنهاد",
        ]
    )
    for item in cohort_summary["students"]:
        analysis.append(
            [
                item["enrollment"],
                _safe_excel_text(item["student_name"]),
                _safe_excel_text(item["student_number"]),
                _safe_excel_text(
                    next(
                        enrollment.class_section.title
                        for enrollment in enrollments
                        if str(enrollment.id) == item["enrollment"]
                    )
                ),
                item["completion_status"],
                item["completion_percent"],
                item["overall_score"],
                item["performance_level"],
                item["first_month"],
                item["last_month"],
                item["change"],
                item["trend_label"],
                item["strongest_domain"]["title"] if item["strongest_domain"] else None,
                item["weakest_domain"]["title"] if item["weakest_domain"] else None,
                item["rank"],
                item["ranked_count"],
                item["recommendation"],
            ]
        )
    _style_sheet(analysis)

    dashboard = workbook.create_sheet("داشبورد")
    dashboard.append(["شاخص", "مقدار", "جزئیات"])
    for key, title in [
        ("students", "تعداد دانش‌آموزان"),
        ("evaluated", "دارای ارزیابی"),
        ("final", "ارزیابی نهایی"),
        ("provisional", "ارزیابی موقت"),
        ("ranked", "رتبه‌بندی‌شده"),
    ]:
        dashboard.append([title, cohort_summary["counts"][key], None])
    dashboard.append([])
    dashboard.append(["روند ماهانه", "میانگین", "تعداد دانش‌آموز"])
    for item in cohort_summary["monthly_trend"]:
        dashboard.append([item["month_no"], item["average"], item["students"]])
    dashboard.append([])
    dashboard.append(["حوزه", "میانگین", None])
    for item in cohort_summary["domain_scores"]:
        dashboard.append([item["title"], item["average"], None])
    _style_sheet(dashboard)

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    output.seek(0)
    return output
