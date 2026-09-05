import math
from datetime import date
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

from .adapters import LoadedImportRows
from .comprehensive import EVALUATION_SHEET

PERSIAN_DIGITS = str.maketrans("۰۱۲۳۴۵۶۷۸۹٠١٢٣٤٥٦٧٨٩", "01234567890123456789")
SOURCE_SUMMARY_KEYS = [
    "educational_score",
    "development_score",
    "character_score",
    "discipline_score",
    "cultural_score",
    "research_score",
    "sport_score",
    "art_score",
    "personal_skills_score",
    "overall_score",
    "performance_level",
    "month_no",
    "completion_ratio",
    "note",
]


def _text(value) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip().translate(PERSIAN_DIGITS)


def _label(value) -> str:
    return "".join(
        _text(value)
        .replace("ي", "ی")
        .replace("ك", "ک")
        .replace("‌", " ")
        .replace("ـ", "")
        .split()
    ).lower()


def _json_scalar(value):
    if value is None or isinstance(value, (str, int, bool)):
        return value
    if isinstance(value, float):
        return value if math.isfinite(value) else str(value)
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def _sheet(workbook, expected_name):
    expected = _label(expected_name)
    return next((item for item in workbook.worksheets if _label(item.title) == expected), None)


def enrich_template_profile(job, loaded: LoadedImportRows):
    """Attach calculated source columns and preserve workbook-level configuration snapshots."""
    evaluation_rows = {
        row.get("__source_row__"): row
        for row in loaded.rows
        if row.get("record_type") == "evaluation" and row.get("__source_row__")
    }
    profile = {"domain_weights": [], "recommendations": []}

    with job.source_file.open("rb") as source:
        payload = source.read()
    workbook = load_workbook(BytesIO(payload), read_only=True, data_only=True)
    try:
        evaluation_sheet = _sheet(workbook, EVALUATION_SHEET)
        if evaluation_sheet is not None and evaluation_rows:
            for source_row, values in enumerate(
                evaluation_sheet.iter_rows(
                    min_row=5,
                    max_row=max(evaluation_rows),
                    min_col=81,
                    max_col=94,
                    values_only=True,
                ),
                start=5,
            ):
                row = evaluation_rows.get(source_row)
                if row is None:
                    continue
                row["__source_summary__"] = {
                    key: _json_scalar(value)
                    for key, value in zip(SOURCE_SUMMARY_KEYS, values, strict=True)
                    if value not in (None, "")
                }

        weights = _sheet(workbook, "تنظیمات وزن‌دهی")
        if weights is not None:
            for values in weights.iter_rows(
                min_row=5, max_row=50, min_col=1, max_col=6, values_only=True
            ):
                code = _text(values[0]).upper()
                if len(code) != 3 or not code.isascii() or not code.isalpha():
                    continue
                profile["domain_weights"].append(
                    {
                        "code": code,
                        "title": _text(values[1]),
                        "weight": _json_scalar(values[2]),
                        "metric_count": _json_scalar(values[3]),
                        "start_column": _text(values[4]),
                        "end_column": _text(values[5]),
                    }
                )

        recommendations = _sheet(workbook, "بانک راهکارها")
        if recommendations is not None:
            for values in recommendations.iter_rows(
                min_row=5, max_row=100, min_col=1, max_col=4, values_only=True
            ):
                if not any(value not in (None, "") for value in values):
                    continue
                profile["recommendations"].append(
                    {
                        "row": _json_scalar(values[0]),
                        "domain": _text(values[1]),
                        "recommendation": _text(values[2]),
                        "usage": _text(values[3]),
                    }
                )
    finally:
        workbook.close()

    return (
        LoadedImportRows(
            rows=loaded.rows,
            source_row_count=loaded.source_row_count,
            adapter=f"{loaded.adapter}-template-aware",
        ),
        profile,
    )
