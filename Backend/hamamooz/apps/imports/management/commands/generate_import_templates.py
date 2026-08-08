from pathlib import Path

from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from hamamooz.apps.imports.models import ImportJob
from hamamooz.apps.imports.services import EXPECTED_HEADERS
from hamamooz.apps.imports.templates import build_smart_evaluation_template
from hamamooz.apps.organizations.models import ClassSection


class Command(BaseCommand):
    help = "Generate the fixed XLSX templates used by the import process."

    def add_arguments(self, parser):
        parser.add_argument(
            "--class-section",
            help="Class-section UUID for a populated smart monthly-evaluation v2 template.",
        )
        parser.add_argument(
            "--output",
            help="Output path for --class-section mode (defaults to docs/import_templates).",
        )

    def handle(self, *args, **options):
        target = Path(settings.BASE_DIR) / "docs" / "import_templates"
        target.mkdir(parents=True, exist_ok=True)
        class_section_id = options.get("class_section")
        if class_section_id:
            try:
                class_section = ClassSection.objects.select_related("school", "academic_year").get(
                    pk=class_section_id
                )
            except (ClassSection.DoesNotExist, ValueError, TypeError) as exc:
                raise CommandError("Class section was not found.") from exc
            default_name = (
                f"smart_monthly_evaluations_{class_section.school.code}_"
                f"{class_section.academic_year.code}_{class_section.code}.xlsx"
            )
            output_path = Path(options.get("output") or target / default_name).resolve()
            output_path.parent.mkdir(parents=True, exist_ok=True)
            template = build_smart_evaluation_template(class_section)
            with output_path.open("wb") as destination:
                destination.write(template.read())
            self.stdout.write(self.style.SUCCESS(str(output_path)))
            return
        examples = {
            ImportJob.ImportType.STUDENTS: ["0012345678", "علی", "احمدی", "2012-09-05", "male"],
            ImportJob.ImportType.ENROLLMENTS: [
                "0012345678",
                "1405-1406",
                "grade-7",
                "7-a",
                "1001",
                "2026-09-23",
            ],
            ImportJob.ImportType.SCORES: [
                "00000000-0000-0000-0000-000000000000",
                "0012345678",
                18.5,
                "present",
                "",
            ],
            ImportJob.ImportType.MONTHLY_EVALUATIONS: [
                "1.0",
                "s1",
                "1405-1406",
                "7-a",
                "101",
                "0012345678",
                4,
                "EDU_01",
                4,
                "",
            ],
        }
        for import_type, headers in EXPECTED_HEADERS.items():
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = import_type
            sheet.append(headers)
            sheet.append(examples[import_type])
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="0A2848")
                cell.alignment = Alignment(horizontal="center")
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            for column in sheet.columns:
                letter = column[0].column_letter
                sheet.column_dimensions[letter].width = max(
                    16, max(len(str(c.value or "")) for c in column) + 2
                )
            path = target / f"{import_type}_template.xlsx"
            workbook.save(path)
            self.stdout.write(self.style.SUCCESS(str(path)))
