from openpyxl import load_workbook


class ImportPreviewService:
    """Read-only analysis phase before committing imported data."""

    def __init__(self, job):
        self.job = job

    def run(self):
        workbook = load_workbook(self.job.source_file, read_only=True, data_only=True)
        sheets = workbook.sheetnames

        return {
            "summary": {
                "students": 0,
                "classes": [],
                "indicators": 0,
                "periods": 0,
                "sheets": sheets,
            },
            "warnings": [],
            "errors": [],
        }
