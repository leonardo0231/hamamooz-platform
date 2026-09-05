from dataclasses import dataclass
from decimal import Decimal, InvalidOperation

from django.conf import settings
from openpyxl.utils import get_column_letter

from hamamooz.apps.evaluations.catalog import (
    FRAMEWORK_VERSION,
    METRIC_CATALOGS,
    metric_catalog_for,
)

from .models import ImportJob

LONG_MONTHLY_HEADERS = [
    "نسخه قالب",
    "کد مدرسه",
    "سال تحصیلی",
    "کد کلاس",
    "شماره دانش‌آموزی",
    "کد ملی",
    "شماره ماه",
    "کد شاخص",
    "امتیاز",
    "توضیحات",
]

LONG_MONTHLY_INTERNAL_HEADERS = [
    "template_version",
    "school_code",
    "academic_year_code",
    "class_code",
    "student_number",
    "national_id",
    "month_no",
    "metric_code",
    "score",
    "note",
]

SMART_TEMPLATE_VERSION = "2.0"
SMART_DATA_SHEET = "ثبت اطلاعات"
SMART_METADATA_SHEET = "__hamamooz_meta"

MONTH_NUMBERS = {
    "تیر": 1,
    "مرداد": 2,
    "شهریور": 3,
    "مهر": 4,
    "آبان": 5,
    "آذر": 6,
    "دی": 7,
    "بهمن": 8,
    "اسفند": 9,
    "فروردین": 10,
    "اردیبهشت": 11,
    "خرداد": 12,
}


@dataclass(frozen=True, slots=True)
class LoadedImportRows:
    rows: list[dict]
    source_row_count: int
    adapter: str


def _text(value) -> str:
    return str(value or "").strip()


def _check_cell_value(value) -> None:
    if isinstance(value, str) and len(value) > 5000:
        raise ValueError("طول مقدار یکی از سلول‌ها بیش از حد مجاز است.")


class LongMonthlyEvaluationAdapter:
    name = "long-monthly-v1"

    @classmethod
    def matches(cls, workbook) -> bool:
        sheet = workbook.active
        header = [_text(cell.value) for cell in sheet[1]]
        return header == LONG_MONTHLY_HEADERS

    @classmethod
    def load(cls, workbook) -> LoadedImportRows:
        sheet = workbook.active
        if sheet.max_column > settings.IMPORT_MAX_COLUMNS:
            raise ValueError("تعداد ستون‌های فایل بیش از حد مجاز است.")
        rows = []
        for source_row, values in enumerate(
            sheet.iter_rows(min_row=2, max_col=len(LONG_MONTHLY_HEADERS), values_only=True),
            start=2,
        ):
            if not any(value not in (None, "") for value in values):
                continue
            if len(rows) >= settings.IMPORT_MAX_ROWS:
                raise ValueError("تعداد ردیف‌های فایل بیش از حد مجاز است.")
            for value in values:
                _check_cell_value(value)
            row = dict(zip(LONG_MONTHLY_INTERNAL_HEADERS, values, strict=True))
            row["framework_version"] = FRAMEWORK_VERSION
            row["__source_row__"] = source_row
            rows.append(row)
        return LoadedImportRows(rows=rows, source_row_count=len(rows), adapter=cls.name)


class SmartWideEvaluationAdapter:
    name = "smart-wide-v2"

    @classmethod
    def matches(cls, workbook) -> bool:
        return (
            SMART_DATA_SHEET in workbook.sheetnames or SMART_METADATA_SHEET in workbook.sheetnames
        )

    @classmethod
    def load(cls, job: ImportJob, workbook) -> LoadedImportRows:
        if SMART_DATA_SHEET not in workbook.sheetnames:
            raise ValueError(f"شیت «{SMART_DATA_SHEET}» در فایل پیدا نشد.")
        if SMART_METADATA_SHEET not in workbook.sheetnames:
            raise ValueError(
                "متادیتای امن قالب هوشمند پیدا نشد؛ فایل را از سامانه و با قالب نسخه ۲ دریافت کنید."
            )

        data_sheet = workbook[SMART_DATA_SHEET]
        metadata_sheet = workbook[SMART_METADATA_SHEET]
        max_columns = getattr(settings, "IMPORT_MAX_SMART_COLUMNS", 100)
        if data_sheet.max_column > max_columns:
            raise ValueError("تعداد ستون‌های قالب هوشمند بیش از حد مجاز است.")

        metadata = {
            _text(row[0]): _text(row[1])
            for row in metadata_sheet.iter_rows(min_col=1, max_col=2, values_only=True)
            if row[0] not in (None, "")
        }
        required_metadata = {
            "template_version",
            "framework_version",
            "school_code",
            "academic_year_code",
            "class_code",
        }
        missing = sorted(required_metadata - metadata.keys())
        if missing:
            raise ValueError(f"متادیتای قالب هوشمند ناقص است: {', '.join(missing)}")
        if metadata["template_version"] != SMART_TEMPLATE_VERSION:
            raise ValueError(f"نسخه قالب هوشمند باید {SMART_TEMPLATE_VERSION} باشد.")
        if metadata["framework_version"] not in METRIC_CATALOGS:
            raise ValueError(f"نسخه چارچوب شاخص‌ها {metadata['framework_version']} پشتیبانی نمی‌شود.")
        catalog = metric_catalog_for(metadata["framework_version"])
        if metadata["school_code"] != job.school.code:
            raise ValueError("کد مدرسه با شعبه انتخاب‌شده در سایت یکسان نیست.")

        enrollments = {}
        for row in metadata_sheet.iter_rows(min_row=2, min_col=4, max_col=7, values_only=True):
            local_code = _text(row[0])
            if not local_code:
                continue
            if local_code in enrollments:
                raise ValueError("کد محلی دانش‌آموز در متادیتای قالب تکراری است.")
            enrollments[local_code] = {
                "enrollment_id": _text(row[1]),
                "national_id": _text(row[2]),
                "student_number": _text(row[3]),
            }
        if not enrollments:
            raise ValueError("هیچ دانش‌آموزی در متادیتای قالب هوشمند ثبت نشده است.")

        metric_columns = {}
        seen_codes = set()
        for row in metadata_sheet.iter_rows(min_row=2, min_col=9, max_col=11, values_only=True):
            if row[0] in (None, ""):
                continue
            try:
                column_index = int(row[0])
            except (TypeError, ValueError) as exc:
                raise ValueError("شماره ستون شاخص در متادیتای قالب معتبر نیست.") from exc
            metric_code = _text(row[1]).upper()
            if metric_code not in catalog:
                raise ValueError("کد شاخص ناشناخته در متادیتای قالب وجود دارد.")
            if column_index in metric_columns or metric_code in seen_codes:
                raise ValueError("نگاشت ستون شاخص در متادیتای قالب تکراری است.")
            expected_title = catalog[metric_code]["title"]
            actual_title = _text(data_sheet.cell(2, column_index).value)
            if actual_title != expected_title or _text(row[2]) != expected_title:
                raise ValueError(
                    f"عنوان ستون {get_column_letter(column_index)} با چارچوب شاخص‌ها یکسان نیست."
                )
            metric_columns[column_index] = metric_code
            seen_codes.add(metric_code)
        if seen_codes != set(catalog):
            raise ValueError(f"نگاشت {len(catalog)} شاخص در متادیتای قالب کامل نیست.")

        expanded = []
        source_row_count = 0
        max_source_rows = getattr(settings, "IMPORT_MAX_SMART_ROWS", settings.IMPORT_MAX_ROWS)
        max_expanded_rows = getattr(settings, "IMPORT_MAX_EXPANDED_ROWS", 100_000)
        note_column = 5 + len(catalog) + 9 + 4
        for source_row, values in enumerate(
            data_sheet.iter_rows(min_row=3, max_col=data_sheet.max_column, values_only=True),
            start=3,
        ):
            local_code = _text(values[2] if len(values) >= 3 else None)
            month_value = values[1] if len(values) >= 2 else None
            entered_metrics = {
                code: values[column_index - 1]
                for column_index, code in metric_columns.items()
                if column_index <= len(values) and values[column_index - 1] not in (None, "")
            }
            note = _text(values[note_column - 1] if note_column <= len(values) else None)
            if not entered_metrics and not note:
                continue
            source_row_count += 1
            if source_row_count > max_source_rows:
                raise ValueError("تعداد ردیف‌های قالب هوشمند بیش از حد مجاز است.")
            if not local_code or month_value in (None, ""):
                expanded.append(
                    {
                        "__source_row__": source_row,
                        "__adapter_error__": "ماه و کد دانش‌آموزی برای هر ردیف الزامی است.",
                    }
                )
                continue
            enrollment = enrollments.get(local_code)
            if enrollment is None:
                expanded.append(
                    {
                        "__source_row__": source_row,
                        "__adapter_error__": "کد دانش‌آموزی در متادیتای امن قالب پیدا نشد.",
                    }
                )
                continue
            month_no = MONTH_NUMBERS.get(_text(month_value))
            if month_no is None:
                try:
                    month_decimal = Decimal(str(month_value))
                    month_no = int(month_decimal)
                    if month_decimal != month_no:
                        raise ValueError
                except (InvalidOperation, TypeError, ValueError):
                    month_no = month_value
            if not entered_metrics:
                expanded.append(
                    {
                        "__source_row__": source_row,
                        "__adapter_error__": "حداقل یک امتیاز شاخص در ردیف الزامی است.",
                    }
                )
                continue
            for metric_code, score in entered_metrics.items():
                _check_cell_value(score)
                expanded.append(
                    {
                        "template_version": SMART_TEMPLATE_VERSION,
                        "framework_version": metadata["framework_version"],
                        "school_code": metadata["school_code"],
                        "academic_year_code": metadata["academic_year_code"],
                        "class_code": metadata["class_code"],
                        "student_number": enrollment["student_number"],
                        "national_id": enrollment["national_id"],
                        "enrollment_id": enrollment["enrollment_id"],
                        "month_no": month_no,
                        "metric_code": metric_code,
                        "score": score,
                        "note": note,
                        "__source_row__": source_row,
                    }
                )
                if len(expanded) > max_expanded_rows:
                    raise ValueError("تعداد امتیازهای استخراج‌شده بیش از حد مجاز است.")
        return LoadedImportRows(
            rows=expanded,
            source_row_count=source_row_count,
            adapter=f"{cls.name}-{metadata['framework_version']}",
        )


def load_monthly_evaluation_xlsx(job: ImportJob, workbook) -> LoadedImportRows:
    if LongMonthlyEvaluationAdapter.matches(workbook):
        return LongMonthlyEvaluationAdapter.load(workbook)
    if SmartWideEvaluationAdapter.matches(workbook):
        return SmartWideEvaluationAdapter.load(job, workbook)
    raise ValueError(
        "فایل نه با قالب طولی فعلی سازگار است و نه متادیتای قالب هوشمند نسخه ۲ را دارد."
    )
