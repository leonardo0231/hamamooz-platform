from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from hamamooz.apps.evaluations.catalog import DOMAIN_DEFINITIONS, FRAMEWORK_VERSION, METRIC_CATALOG

from .comprehensive import (
    CLASS_HEADERS,
    CLASS_SHEET,
    EVALUATION_IDENTITY_HEADERS,
    EVALUATION_SHEET,
    STUDENT_HEADERS,
    STUDENT_SHEET,
)

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
CALC_FILL = PatternFill("solid", fgColor="E7E6E6")
WHITE_BOLD = Font(name="Vazirmatn", bold=True, color="FFFFFF")
BODY_FONT = Font(name="Vazirmatn", size=10)
MONTH_LABELS = [
    "تیر",
    "مرداد",
    "شهریور",
    "مهر",
    "آبان",
    "آذر",
    "دی",
    "بهمن",
    "اسفند",
    "فروردین",
    "اردیبهشت",
    "خرداد",
]


def _style_header(row):
    for cell in row:
        cell.fill = HEADER_FILL
        cell.font = WHITE_BOLD
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _metric_bounds_by_domain():
    result = {}
    for domain_code in DOMAIN_DEFINITIONS:
        columns = [
            7 + index
            for index, code in enumerate(METRIC_CATALOG)
            if METRIC_CATALOG[code]["domain_code"] == domain_code
        ]
        result[domain_code] = (min(columns), max(columns))
    return result


def build_comprehensive_school_template() -> BytesIO:
    """Generate the official comprehensive-school workbook from the active catalog.

    The import contract is intentionally generated from the same catalog used by
    validation so metric-count or title changes cannot leave the downloadable
    template behind the backend implementation.
    """

    workbook = Workbook()
    guide = workbook.active
    guide.title = "راهنما"
    guide.sheet_view.rightToLeft = True
    guide.append(["قالب جامع مدرسه هم‌آموز"])
    guide.append(["نسخه چارچوب شاخص‌ها", FRAMEWORK_VERSION])
    guide.append(
        [
            "راهنما",
            "ابتدا کلاس‌ها و دانش‌آموزان را تکمیل کنید؛ سپس امتیازهای ۰ تا ۵ را در شیت ثبت اطلاعات وارد کنید.",
        ]
    )
    guide.append(["نکته", "کد ملی باید دقیقاً ۱۰ رقم و تاریخ تولد باید با قالب YYYY-MM-DD ثبت شود."])
    guide.column_dimensions["A"].width = 28
    guide.column_dimensions["B"].width = 90

    classes = workbook.create_sheet(CLASS_SHEET)
    classes.sheet_view.rightToLeft = True
    classes.append(["کلاس‌بندی"])
    classes.append(["اطلاعات هر کلاس را در یک ردیف وارد کنید."])
    classes.append([])
    classes.append(CLASS_HEADERS + ["تعداد ثبت‌شده", "وضعیت"])
    _style_header(classes[4])
    for index in range(1, 31):
        classes.append([index, None, None, None, None, None, None, None, None])
    classes.freeze_panes = "A5"
    for column, width in {
        "A": 10,
        "B": 16,
        "C": 16,
        "D": 14,
        "E": 26,
        "F": 18,
        "G": 12,
        "H": 16,
        "I": 14,
    }.items():
        classes.column_dimensions[column].width = width

    students = workbook.create_sheet(STUDENT_SHEET)
    students.sheet_view.rightToLeft = True
    students.append(["فهرست دانش‌آموزان"])
    students.append(["اطلاعات هویتی و کلاس هر دانش‌آموز را وارد کنید."])
    students.append([])
    students.append(STUDENT_HEADERS + ["نام کلاس", "پایه", "وضعیت اطلاعات"])
    _style_header(students[4])
    for index in range(1, 101):
        students.append([index, index, None, None, None, None, None, None, None, None, None, None])
    students.freeze_panes = "A5"
    for column, width in {
        "A": 10,
        "B": 12,
        "C": 16,
        "D": 20,
        "E": 18,
        "F": 22,
        "G": 12,
        "H": 16,
        "I": 14,
        "J": 26,
        "K": 14,
        "L": 16,
    }.items():
        students.column_dimensions[column].width = width

    weights = workbook.create_sheet("تنظیمات وزن‌دهی")
    weights.sheet_view.rightToLeft = True
    weights.append(["تنظیمات وزن‌دهی حوزه‌ها"])
    weights.append(["اوزان رسمی سامانه؛ نام و کد حوزه‌ها را تغییر ندهید."])
    weights.append([])
    weights.append(
        ["کد حوزه", "نام حوزه", "وزن درصدی", "تعداد شاخص", "شروع ستون شاخص", "پایان ستون شاخص"]
    )
    _style_header(weights[4])
    domain_bounds = _metric_bounds_by_domain()
    for domain_code, (title, weight) in DOMAIN_DEFINITIONS.items():
        start_column, end_column = domain_bounds[domain_code]
        metric_count = sum(
            definition["domain_code"] == domain_code for definition in METRIC_CATALOG.values()
        )
        weights.append(
            [
                domain_code,
                title,
                weight,
                metric_count,
                get_column_letter(start_column),
                get_column_letter(end_column),
            ]
        )
    weights.append([])
    weights.append(["جمع وزن‌ها", None, "=SUM(C5:C13)", None, None, None])
    for column, width in {"A": 14, "B": 24, "C": 14, "D": 14, "E": 18, "F": 18}.items():
        weights.column_dimensions[column].width = width

    data = workbook.create_sheet(EVALUATION_SHEET)
    data.sheet_view.rightToLeft = True
    metric_codes = list(METRIC_CATALOG)
    calculated_headers = [
        "امتیاز آموزشی (۰ تا ۲۰)",
        "امتیاز پرورشی (۰ تا ۲۰)",
        "امتیاز تربیتی (۰ تا ۲۰)",
        "امتیاز انضباطی (۰ تا ۲۰)",
        "امتیاز فرهنگی (۰ تا ۲۰)",
        "امتیاز پژوهشی (۰ تا ۲۰)",
        "امتیاز ورزشی (۰ تا ۲۰)",
        "امتیاز هنری (۰ تا ۲۰)",
        "امتیاز مهارت‌های فردی (۰ تا ۲۰)",
        "میانگین وزنی نهایی",
        "سطح عملکرد",
        "شماره ماه",
        "درصد تکمیل",
        "توضیحات",
    ]
    data.append(["ثبت اطلاعات ماهانه"])
    data.append([f"چارچوب {FRAMEWORK_VERSION} — {len(metric_codes)} شاخص"])
    data.append([])
    headers = (
        EVALUATION_IDENTITY_HEADERS
        + [f"{code} | {METRIC_CATALOG[code]['title']}" for code in metric_codes]
        + [None] * (80 - (6 + len(metric_codes)))
        + calculated_headers
    )
    data.append(headers)
    _style_header(data[4])

    score_validation = DataValidation(type="whole", operator="between", formula1="0", formula2="5")
    score_validation.error = "امتیاز باید عدد صحیح ۰ تا ۵ باشد."
    score_validation.errorTitle = "امتیاز نامعتبر"
    score_validation.showErrorMessage = True
    data.add_data_validation(score_validation)

    metric_start = 7
    metric_end = metric_start + len(metric_codes) - 1
    metric_start_letter = get_column_letter(metric_start)
    metric_end_letter = get_column_letter(metric_end)
    domain_score_start = 81
    domain_score_end = 89
    overall_column = 90
    level_column = 91
    month_number_column = 92
    completion_column = 93
    note_column = 94

    row_number = 5
    for student_index in range(1, 101):
        for month_index, month_label in enumerate(MONTH_LABELS, start=1):
            data.cell(row_number, 1, row_number - 4)
            data.cell(row_number, 2, month_label)
            data.cell(row_number, 3, student_index)
            data.cell(row_number, month_number_column, month_index)
            for domain_offset, domain_code in enumerate(DOMAIN_DEFINITIONS):
                first, last = domain_bounds[domain_code]
                first_letter = get_column_letter(first)
                last_letter = get_column_letter(last)
                target = get_column_letter(domain_score_start + domain_offset)
                data[f"{target}{row_number}"] = (
                    f'=IF(COUNT({first_letter}{row_number}:{last_letter}{row_number})=0,"",'
                    f"AVERAGE({first_letter}{row_number}:{last_letter}{row_number})*4)"
                )
            overall_letter = get_column_letter(overall_column)
            domain_start_letter = get_column_letter(domain_score_start)
            domain_end_letter = get_column_letter(domain_score_end)
            data[f"{overall_letter}{row_number}"] = (
                f'=IF(COUNT({metric_start_letter}{row_number}:{metric_end_letter}{row_number})<{len(metric_codes)},"",'
                f"SUMPRODUCT({domain_start_letter}{row_number}:{domain_end_letter}{row_number},"
                "'تنظیمات وزن‌دهی'!$C$5:$C$13)/100)"
            )
            level_letter = get_column_letter(level_column)
            data[f"{level_letter}{row_number}"] = (
                f'=IF({overall_letter}{row_number}="","",IF({overall_letter}{row_number}>=17,"عالی",'
                f'IF({overall_letter}{row_number}>=14,"خوب",IF({overall_letter}{row_number}>=10,"متوسط","ضعیف"))))'
            )
            completion_letter = get_column_letter(completion_column)
            data[f"{completion_letter}{row_number}"] = (
                f"=COUNT({metric_start_letter}{row_number}:{metric_end_letter}{row_number})/{len(metric_codes)}"
            )
            row_number += 1

    last_row = row_number - 1
    score_validation.add(f"{metric_start_letter}5:{metric_end_letter}{last_row}")
    data.freeze_panes = "G5"
    data.auto_filter.ref = f"A4:{get_column_letter(note_column)}{last_row}"
    for column_index in range(1, note_column + 1):
        data.column_dimensions[get_column_letter(column_index)].width = 13
    data.column_dimensions["E"].width = 30
    data.column_dimensions[get_column_letter(note_column)].width = 36
    data.column_dimensions[get_column_letter(completion_column)].width = 16

    for row in data.iter_rows(min_row=5, max_row=last_row, min_col=1, max_col=note_column):
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.protection = Protection(locked=True)
        for cell in row[metric_start - 1 : metric_end]:
            cell.fill = INPUT_FILL
            cell.protection = Protection(locked=False)
        row[note_column - 1].fill = INPUT_FILL
        row[note_column - 1].protection = Protection(locked=False)
        for cell in row[domain_score_start - 1 : completion_column]:
            cell.fill = CALC_FILL

    data.protection.sheet = True
    data.protection.password = "hamamooz-v2"

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    output.seek(0)
    return output
