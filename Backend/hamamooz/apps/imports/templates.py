from io import BytesIO
from math import ceil

from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from hamamooz.apps.evaluations.catalog import (
    DOMAIN_DEFINITIONS,
    FRAMEWORK_VERSION,
    METRIC_CATALOG,
)
from hamamooz.apps.students.models import Enrollment

from .adapters import MONTH_NUMBERS, SMART_METADATA_SHEET, SMART_TEMPLATE_VERSION

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
CALC_FILL = PatternFill("solid", fgColor="E7E6E6")
WHITE_BOLD = Font(name="Vazirmatn", bold=True, color="FFFFFF")
BODY_FONT = Font(name="Vazirmatn", size=10)


def _safe_excel_text(value):
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value


def _set_plain_text(cell, value) -> None:
    cell.value = str(value)
    cell.data_type = "s"


def _style_header(row) -> None:
    for cell in row:
        cell.fill = HEADER_FILL
        cell.font = WHITE_BOLD
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def build_smart_evaluation_template(class_section) -> BytesIO:
    enrollments = list(
        Enrollment.objects.filter(
            class_section=class_section,
            school=class_section.school,
            academic_year=class_section.academic_year,
            status=Enrollment.Status.ACTIVE,
        )
        .select_related("student", "class_section")
        .order_by("student__last_name", "student__first_name", "student_number")
    )
    workbook = Workbook()
    guide = workbook.active
    guide.title = "راهنما"
    guide.sheet_view.rightToLeft = True
    guide.append(["قالب هوشمند ارزیابی ماهانه هم‌آموز — نسخه ۲"])
    guide.append(["مدرسه", _safe_excel_text(class_section.school.name)])
    guide.append(["سال تحصیلی", _safe_excel_text(class_section.academic_year.title)])
    guide.append(["کلاس", _safe_excel_text(class_section.title)])
    guide.append(["راهنما", "فقط سلول‌های زرد امتیاز ۰ تا ۵ و توضیحات را ویرایش کنید."])
    guide.append(
        [
            "وضعیت تکمیل",
            "سطح عملکرد و نمره نهایی فقط پس از رسیدن به حد تکمیل تعیین‌شده نمایش داده می‌شود.",
        ]
    )
    guide.column_dimensions["A"].width = 20
    guide.column_dimensions["B"].width = 80

    weights = workbook.create_sheet("تنظیمات وزن‌دهی")
    weights.sheet_view.rightToLeft = True
    weights.append(["کد حوزه", "حیطه ارزیابی", "وزن (٪)", "تعداد زیرمعیار"])
    _style_header(weights[1])
    for domain_code, (title, weight) in DOMAIN_DEFINITIONS.items():
        metric_count = sum(
            definition["domain_code"] == domain_code for definition in METRIC_CATALOG.values()
        )
        weights.append([domain_code, title, weight, metric_count])
    weights.append(["", "جمع", "=SUM(C2:C10)", "=SUM(D2:D10)"])
    weights.column_dimensions["A"].width = 14
    weights.column_dimensions["B"].width = 24
    weights.column_dimensions["C"].width = 14
    weights.column_dimensions["D"].width = 18

    students = workbook.create_sheet("دانش آموزان")
    students.sheet_view.rightToLeft = True
    students.append(["ردیف", "کد دانش‌آموزی", "نام و نام خانوادگی", "کلاس", "شماره دانش‌آموزی"])
    _style_header(students[1])
    local_codes = {}
    for index, enrollment in enumerate(enrollments, start=1):
        local_code = str(index)
        local_codes[enrollment.id] = local_code
        students.append(
            [
                index,
                local_code,
                _safe_excel_text(enrollment.student.full_name),
                _safe_excel_text(class_section.title),
                _safe_excel_text(enrollment.student_number),
            ]
        )
    students.freeze_panes = "A2"
    for column, width in {"A": 10, "B": 18, "C": 32, "D": 24, "E": 22}.items():
        students.column_dimensions[column].width = width

    data = workbook.create_sheet("ثبت اطلاعات")
    data.sheet_view.rightToLeft = True
    metric_codes = list(METRIC_CATALOG)
    headers = (
        ["ردیف", "ماه", "کد دانش‌آموزی", "نام و نام خانوادگی", "کلاس"]
        + [METRIC_CATALOG[code]["title"] for code in metric_codes]
        + [f"میانگین {title} (۲۰)" for title, _weight in DOMAIN_DEFINITIONS.values()]
        + ["نمره وزنی نهایی (۲۰)", "سطح عملکرد", "شماره ماه", "توضیحات"]
    )
    data.append([None] * len(headers))
    data.append(headers)
    _style_header(data[2])

    metric_column_by_code = {code: index + 6 for index, code in enumerate(metric_codes)}
    for domain_code, (title, _weight) in DOMAIN_DEFINITIONS.items():
        domain_columns = [
            metric_column_by_code[code]
            for code in metric_codes
            if METRIC_CATALOG[code]["domain_code"] == domain_code
        ]
        start = get_column_letter(min(domain_columns))
        end = get_column_letter(max(domain_columns))
        data.merge_cells(f"{start}1:{end}1")
        cell = data[f"{start}1"]
        cell.value = f"{title} (۰ تا ۵)"
        cell.fill = HEADER_FILL
        cell.font = WHITE_BOLD
        cell.alignment = Alignment(horizontal="center")

    threshold_percent = max(
        1, min(100, getattr(settings, "EVALUATION_FINAL_COMPLETION_PERCENT", 100))
    )
    required_count = ceil(len(METRIC_CATALOG) * threshold_percent / 100)
    metric_start = get_column_letter(6)
    metric_end = get_column_letter(5 + len(metric_codes))
    domain_score_start = 6 + len(metric_codes)
    month_labels = list(MONTH_NUMBERS)
    row_number = 3
    for enrollment in enrollments:
        for month_no, month_label in enumerate(month_labels, start=1):
            values = [
                row_number - 2,
                month_label,
                local_codes[enrollment.id],
                _safe_excel_text(enrollment.student.full_name),
                _safe_excel_text(class_section.title),
            ] + [None] * (len(headers) - 5)
            data.append(values)
            for domain_offset, domain_code in enumerate(DOMAIN_DEFINITIONS):
                domain_columns = [
                    metric_column_by_code[code]
                    for code in metric_codes
                    if METRIC_CATALOG[code]["domain_code"] == domain_code
                ]
                first = get_column_letter(min(domain_columns))
                last = get_column_letter(max(domain_columns))
                target = get_column_letter(domain_score_start + domain_offset)
                data[f"{target}{row_number}"] = (
                    f'=IF(COUNT({first}{row_number}:{last}{row_number})=0,"",'
                    f"AVERAGE({first}{row_number}:{last}{row_number})*4)"
                )
            overall_column = get_column_letter(domain_score_start + len(DOMAIN_DEFINITIONS))
            level_column = get_column_letter(domain_score_start + len(DOMAIN_DEFINITIONS) + 1)
            domain_end = get_column_letter(domain_score_start + len(DOMAIN_DEFINITIONS) - 1)
            data[f"{overall_column}{row_number}"] = (
                f'=IF(COUNT({metric_start}{row_number}:{metric_end}{row_number})<{required_count},"",'
                f"SUMPRODUCT({get_column_letter(domain_score_start)}{row_number}:{domain_end}{row_number},"
                "'تنظیمات وزن‌دهی'!$C$2:$C$10)/100)"
            )
            data[f"{level_column}{row_number}"] = (
                f'=IF({overall_column}{row_number}="","",IF({overall_column}{row_number}>=17,"عالی",'
                f'IF({overall_column}{row_number}>=14,"خوب",IF({overall_column}{row_number}>=10,'
                '"متوسط","ضعیف"))))'
            )
            data.cell(row_number, len(headers) - 1, month_no)
            row_number += 1

    last_row = max(3, row_number - 1)
    score_validation = DataValidation(type="whole", operator="between", formula1="0", formula2="5")
    score_validation.error = "امتیاز باید عدد صحیح ۰ تا ۵ باشد."
    score_validation.errorTitle = "امتیاز نامعتبر"
    score_validation.showErrorMessage = True
    data.add_data_validation(score_validation)
    score_validation.add(f"{metric_start}3:{metric_end}{last_row}")
    data.freeze_panes = "F3"
    data.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{last_row}"
    data.row_dimensions[1].height = 24
    data.row_dimensions[2].height = 42
    for cell in data[2]:
        cell.protection = Protection(locked=True)
    for row in data.iter_rows(min_row=3, max_row=last_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.protection = Protection(locked=True)
        for cell in row[5 : 5 + len(metric_codes)]:
            cell.fill = INPUT_FILL
            cell.protection = Protection(locked=False)
        for cell in row[5 + len(metric_codes) : -1]:
            cell.fill = CALC_FILL
        row[-1].fill = INPUT_FILL
        row[-1].protection = Protection(locked=False)
    data.protection.sheet = True
    data.protection.password = "hamamooz-v2"
    for index in range(1, len(headers) + 1):
        data.column_dimensions[get_column_letter(index)].width = 13
    for index in [1, 2, 3]:
        data.column_dimensions[get_column_letter(index)].width = 12
    data.column_dimensions["D"].width = 28
    data.column_dimensions["E"].width = 22
    data.column_dimensions[get_column_letter(len(headers))].width = 36

    metadata = workbook.create_sheet(SMART_METADATA_SHEET)
    metadata.append(["template_version", SMART_TEMPLATE_VERSION])
    metadata.append(["framework_version", FRAMEWORK_VERSION])
    metadata.append(["school_code", None])
    metadata.append(["academic_year_code", None])
    metadata.append(["class_code", None])
    _set_plain_text(metadata["B3"], class_section.school.code)
    _set_plain_text(metadata["B4"], class_section.academic_year.code)
    _set_plain_text(metadata["B5"], class_section.code)
    metadata["D1"] = "local_student_code"
    metadata["E1"] = "enrollment_id"
    metadata["F1"] = "national_id"
    metadata["G1"] = "student_number"
    for index, enrollment in enumerate(enrollments, start=2):
        _set_plain_text(metadata.cell(index, 4), local_codes[enrollment.id])
        _set_plain_text(metadata.cell(index, 5), enrollment.id)
        _set_plain_text(metadata.cell(index, 6), enrollment.student.national_id)
        _set_plain_text(metadata.cell(index, 7), enrollment.student_number)
    metadata["I1"] = "excel_column"
    metadata["J1"] = "metric_code"
    metadata["K1"] = "metric_title"
    for index, code in enumerate(metric_codes, start=2):
        metadata.cell(index, 9, metric_column_by_code[code])
        metadata.cell(index, 10, code)
        metadata.cell(index, 11, METRIC_CATALOG[code]["title"])
    metadata.protection.sheet = True
    metadata.protection.password = "hamamooz-v2"
    metadata.sheet_state = "veryHidden"

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    output.seek(0)
    return output
