from io import BytesIO

from openpyxl import load_workbook

from hamamooz.apps.evaluations.catalog import FRAMEWORK_VERSION
from hamamooz.apps.evaluations.models import MonthlyEvaluation
from hamamooz.apps.organizations.models import ClassSection
from hamamooz.apps.students.models import Enrollment, Student

from .adapters import LoadedImportRows
from .comprehensive import (
    COMPREHENSIVE_TEMPLATE_VERSION,
    EVALUATION_SHEET,
    apply_comprehensive_workbook,
    validate_comprehensive_workbook,
)

PERSIAN_DIGITS = str.maketrans("۰۱۲۳۴۵۶۷۸۹٠١٢٣٤٥٦٧٨٩", "01234567890123456789")


def _text(value) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip().translate(PERSIAN_DIGITS)


def _label(value) -> str:
    return "".join(_text(value).replace("ي", "ی").replace("ك", "ک").replace("‌", " ").split())


def _strict_national_id(value) -> str:
    national_id = _text(value)
    if len(national_id) != 10 or not national_id.isdigit():
        raise ValueError(
            "کد ملی باید دقیقاً ۱۰ رقم باشد؛ صفر ابتدای کد ملی را حذف نکنید "
            "و ستون را Text نگه دارید."
        )
    return national_id


def _find_sheet(workbook, expected_name):
    expected = _label(expected_name)
    matches = [sheet for sheet in workbook.worksheets if _label(sheet.title) == expected]
    if len(matches) != 1:
        raise ValueError(f"شیت «{expected_name}» در فایل پیدا نشد یا بیش از یک‌بار وجود دارد.")
    return matches[0]


def enrich_comprehensive_rows(job, loaded: LoadedImportRows) -> LoadedImportRows:
    """Attach the visible identity columns from the evaluation sheet to parsed rows.

    The original parser intentionally keys evaluation rows by the workbook-local code. For
    safety we also retain national-id, full-name and class-code cells so validation can prove
    that formulas/edited cells still point at the same student before any database write.
    """

    evaluation_rows = {
        row["__source_row__"]: row for row in loaded.rows if row.get("record_type") == "evaluation"
    }
    if not evaluation_rows:
        return loaded

    with job.source_file.open("rb") as source:
        payload = source.read()
    workbook = load_workbook(BytesIO(payload), read_only=True, data_only=True)
    try:
        sheet = _find_sheet(workbook, EVALUATION_SHEET)
        for source_row, values in enumerate(
            sheet.iter_rows(min_row=5, max_row=1204, min_col=1, max_col=6, values_only=True),
            start=5,
        ):
            row = evaluation_rows.get(source_row)
            if row is None:
                continue
            row["__evaluation_national_id__"] = values[3]
            row["__evaluation_full_name__"] = values[4]
            row["__evaluation_class_code__"] = values[5]
    finally:
        workbook.close()

    return LoadedImportRows(
        rows=loaded.rows,
        source_row_count=loaded.source_row_count,
        adapter="comprehensive-school-v1-hardened",
    )


def _identity_error(row, column, code, message):
    return {
        "sheet": row.get("__sheet__", ""),
        "row": row.get("__source_row__"),
        "column": column,
        "code": code,
        "message": message,
    }


def validate_hardened_comprehensive_workbook(job, rows):
    identity_errors = []
    students_by_local_code = {}

    for row in rows:
        if row.get("record_type") != "student":
            continue
        local_code = _text(row.get("local_code"))
        try:
            national_id = _strict_national_id(row.get("national_id"))
        except ValueError as exc:
            identity_errors.append(_identity_error(row, "کد ملی", "national_id_format", str(exc)))
            continue
        students_by_local_code[local_code] = {
            "national_id": national_id,
            "class_code": _text(row.get("class_code")),
            "full_name": f"{_text(row.get('first_name'))} {_text(row.get('last_name'))}".strip(),
        }

    for row in rows:
        if row.get("record_type") != "evaluation":
            continue
        local_code = _text(row.get("local_code"))
        student = students_by_local_code.get(local_code)
        if student is None:
            continue

        evaluation_national_id = _text(row.get("__evaluation_national_id__"))
        if evaluation_national_id:
            try:
                evaluation_national_id = _strict_national_id(evaluation_national_id)
            except ValueError as exc:
                identity_errors.append(
                    _identity_error(row, "کد ملی", "evaluation_national_id_format", str(exc))
                )
            else:
                if evaluation_national_id != student["national_id"]:
                    identity_errors.append(
                        _identity_error(
                            row,
                            "کد ملی",
                            "evaluation_identity_mismatch",
                            "کد ملی ردیف ارزیابی با دانش‌آموز متناظر در شیت "
                            "«دانش‌آموزان» یکسان نیست.",
                        )
                    )

        evaluation_class_code = _text(row.get("__evaluation_class_code__"))
        if evaluation_class_code and evaluation_class_code != student["class_code"]:
            identity_errors.append(
                _identity_error(
                    row,
                    "کد کلاس",
                    "evaluation_class_mismatch",
                    "کد کلاس ردیف ارزیابی با کلاس دانش‌آموز در شیت «دانش‌آموزان» یکسان نیست.",
                )
            )

        evaluation_full_name = _text(row.get("__evaluation_full_name__"))
        if evaluation_full_name and _label(evaluation_full_name) != _label(student["full_name"]):
            identity_errors.append(
                _identity_error(
                    row,
                    "نام و نام خانوادگی",
                    "evaluation_name_mismatch",
                    "نام ردیف ارزیابی با دانش‌آموز متناظر در شیت «دانش‌آموزان» یکسان نیست.",
                )
            )

    prepared, base_errors = validate_comprehensive_workbook(job, rows)
    return prepared, [*identity_errors, *base_errors]


def _snapshot_change_counts(job, prepared):
    academic_year = prepared["academic_year"]
    if academic_year is None:
        return {}

    class_codes = [item["code"] for item in prepared["classes"]]
    existing_classes = {
        item.code: item
        for item in ClassSection.all_objects.filter(
            school=job.school,
            academic_year=academic_year,
            code__in=class_codes,
        )
    }
    class_unchanged = 0
    for item in prepared["classes"]:
        existing = existing_classes.get(item["code"])
        if (
            existing is not None
            and not existing.is_deleted
            and existing.is_active
            and existing.title == item["title"]
            and existing.grade_level_id == item["grade"].id
            and existing.capacity == item["capacity"]
        ):
            class_unchanged += 1

    national_ids = [item["national_id"] for item in prepared["students"]]
    existing_students = {
        item.national_id: item
        for item in Student.all_objects.filter(
            organization=job.organization,
            national_id__in=national_ids,
        )
    }
    student_unchanged = 0
    for item in prepared["students"]:
        existing = existing_students.get(item["national_id"])
        if (
            existing is not None
            and not existing.is_deleted
            and existing.status == Student.Status.ACTIVE
            and existing.first_name == item["first_name"]
            and existing.last_name == item["last_name"]
            and existing.birth_date == item["birth_date"]
            and existing.gender == item["gender"]
        ):
            student_unchanged += 1

    existing_enrollments = {}
    student_ids = [student.id for student in existing_students.values()]
    for enrollment in (
        Enrollment.all_objects.filter(student_id__in=student_ids, academic_year=academic_year)
        .select_related("student", "class_section")
        .order_by("student_id", "-updated_at")
    ):
        existing_enrollments.setdefault(enrollment.student.national_id, enrollment)

    prepared_classes = {item["code"]: item for item in prepared["classes"]}
    enrollment_unchanged = 0
    evaluation_unchanged = 0
    metric_created = 0
    metric_updated = 0
    metric_unchanged = 0

    evaluation_map = {}
    enrollment_ids = [item.id for item in existing_enrollments.values()]
    for evaluation in MonthlyEvaluation.objects.filter(
        enrollment_id__in=enrollment_ids,
        framework_version=FRAMEWORK_VERSION,
    ).prefetch_related("metric_scores"):
        evaluation_map[(evaluation.enrollment_id, evaluation.month_no)] = evaluation

    prepared_students_by_local = {item["local_code"]: item for item in prepared["students"]}
    for item in prepared["students"]:
        enrollment = existing_enrollments.get(item["national_id"])
        class_item = prepared_classes[item["class_code"]]
        existing_class = existing_classes.get(item["class_code"])
        if (
            enrollment is not None
            and existing_class is not None
            and not enrollment.is_deleted
            and enrollment.status == Enrollment.Status.ACTIVE
            and enrollment.school_id == job.school_id
            and enrollment.grade_level_id == class_item["grade"].id
            and enrollment.class_section_id == existing_class.id
            and enrollment.student_number == item["student_number"]
            and enrollment.enrolled_on == academic_year.starts_on
            and enrollment.left_on is None
        ):
            enrollment_unchanged += 1

    for item in prepared["evaluations"]:
        student_item = prepared_students_by_local[item["local_code"]]
        enrollment = existing_enrollments.get(student_item["national_id"])
        if enrollment is None:
            metric_created += len(item["metrics"])
            continue
        evaluation = evaluation_map.get((enrollment.id, item["month_no"]))
        if evaluation is None:
            metric_created += len(item["metrics"])
            continue
        existing_metrics = {
            score.metric_code: score.value for score in evaluation.metric_scores.all()
        }
        metrics_same = True
        for code, value in item["metrics"].items():
            if code not in existing_metrics:
                metric_created += 1
                metrics_same = False
            elif existing_metrics[code] == value:
                metric_unchanged += 1
            else:
                metric_updated += 1
                metrics_same = False
        if evaluation.note == item["note"] and metrics_same:
            evaluation_unchanged += 1

    return {
        "classes_unchanged": class_unchanged,
        "students_unchanged": student_unchanged,
        "enrollments_unchanged": enrollment_unchanged,
        "evaluations_unchanged": evaluation_unchanged,
        "metric_scores_created": metric_created,
        "metric_scores_updated": metric_updated,
        "metric_scores_unchanged": metric_unchanged,
    }


def apply_hardened_comprehensive_workbook(job, prepared):
    snapshot = _snapshot_change_counts(job, prepared)
    summary = apply_comprehensive_workbook(job, prepared)

    summary["classes_unchanged"] = snapshot.get("classes_unchanged", 0)
    summary["students_unchanged"] = snapshot.get("students_unchanged", 0)
    summary["enrollments_unchanged"] = snapshot.get("enrollments_unchanged", 0)
    summary["evaluations_unchanged"] = snapshot.get("evaluations_unchanged", 0)

    summary["classes_updated"] = max(
        0, summary.get("classes_updated", 0) - summary["classes_unchanged"]
    )
    summary["students_updated"] = max(
        0, summary.get("students_updated", 0) - summary["students_unchanged"]
    )
    summary["enrollments_updated"] = max(
        0, summary.get("enrollments_updated", 0) - summary["enrollments_unchanged"]
    )
    summary["evaluations_updated"] = max(
        0, summary.get("evaluations_updated", 0) - summary["evaluations_unchanged"]
    )

    summary.update(
        {
            "metric_scores_created": snapshot.get("metric_scores_created", 0),
            "metric_scores_updated": snapshot.get("metric_scores_updated", 0),
            "metric_scores_unchanged": snapshot.get("metric_scores_unchanged", 0),
            "records_deleted": 0,
            "delete_policy": "explicit_manual_only",
            "template_version": COMPREHENSIVE_TEMPLATE_VERSION,
            "source": "comprehensive_school",
            "classes_total": len(prepared["classes"]),
            "students_total": len(prepared["students"]),
            "evaluations_total": len(prepared["evaluations"]),
        }
    )
    return summary
