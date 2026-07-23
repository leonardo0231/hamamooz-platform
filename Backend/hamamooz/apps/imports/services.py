from collections import defaultdict
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from io import BytesIO
from zipfile import BadZipFile, ZipFile

from django.conf import settings
from django.core.exceptions import ValidationError as DjangoValidationError
from django.db import transaction
from django.db.models import Count
from django.utils import timezone
from openpyxl import load_workbook

from hamamooz.apps.academics.models import Assessment, Score
from hamamooz.apps.academics.services import bulk_upsert_scores
from hamamooz.apps.organizations.models import AcademicYear, ClassSection, GradeLevel
from hamamooz.apps.students.models import Enrollment, Student

from .models import ImportJob

EXPECTED_HEADERS = {
    ImportJob.ImportType.STUDENTS: [
        "national_id",
        "first_name",
        "last_name",
        "birth_date",
        "gender",
    ],
    ImportJob.ImportType.ENROLLMENTS: [
        "national_id",
        "academic_year_code",
        "grade_code",
        "class_code",
        "student_number",
        "enrolled_on",
    ],
    ImportJob.ImportType.SCORES: [
        "assessment_id",
        "national_id",
        "value",
        "status",
        "note",
    ],
}


def _as_date(value, field):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return date.fromisoformat(str(value).strip())
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{field} باید با قالب YYYY-MM-DD باشد.") from exc


def _as_decimal(value):
    if value in (None, ""):
        return None
    try:
        return Decimal(str(value))
    except InvalidOperation as exc:
        raise ValueError("value باید عددی باشد.") from exc


def _model_errors(exc):
    if hasattr(exc, "message_dict"):
        return "; ".join(
            f"{key}: {', '.join(map(str, messages))}" for key, messages in exc.message_dict.items()
        )
    return "; ".join(exc.messages)


def _load_rows(job):
    with job.source_file.open("rb") as source:
        payload = source.read()
    try:
        with ZipFile(BytesIO(payload)) as archive:
            infos = archive.infolist()
            total_uncompressed = sum(item.file_size for item in infos)
            if total_uncompressed > settings.IMPORT_MAX_UNCOMPRESSED_BYTES:
                raise ValueError("حجم بازشده فایل XLSX بیش از حد مجاز است.")
            for item in infos:
                if item.file_size and item.compress_size:
                    ratio = item.file_size / item.compress_size
                    if ratio > 200:
                        raise ValueError("نسبت فشرده‌سازی فایل XLSX غیرعادی است.")
    except BadZipFile:
        raise

    workbook = load_workbook(BytesIO(payload), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        if sheet.max_column > settings.IMPORT_MAX_COLUMNS:
            raise ValueError("تعداد ستون‌های فایل بیش از حد مجاز است.")
        rows = sheet.iter_rows(values_only=True)
        try:
            header = [str(value).strip() if value is not None else "" for value in next(rows)]
        except StopIteration as exc:
            raise ValueError("فایل خالی است.") from exc
        expected = EXPECTED_HEADERS[job.import_type]
        if header != expected:
            raise ValueError(f"ستون‌ها باید دقیقاً به این ترتیب باشند: {', '.join(expected)}")
        loaded = []
        for values in rows:
            if not any(v not in (None, "") for v in values):
                continue
            if len(loaded) >= settings.IMPORT_MAX_ROWS:
                raise ValueError("تعداد ردیف‌های فایل بیش از حد مجاز است.")
            normalized = []
            for value in values:
                if isinstance(value, str) and len(value) > 5000:
                    raise ValueError("طول مقدار یکی از سلول‌ها بیش از حد مجاز است.")
                normalized.append(value)
            loaded.append(dict(zip(header, normalized, strict=True)))
        return loaded
    finally:
        workbook.close()


def _validate_students(job, rows):
    prepared, errors, seen = [], [], set()
    existing = set(
        Student.all_objects.filter(organization=job.organization).values_list(
            "national_id", flat=True
        )
    )
    for number, row in enumerate(rows, start=2):
        try:
            national_id = str(row["national_id"] or "").strip().zfill(10)
            if national_id in seen or national_id in existing:
                raise ValueError("کد ملی تکراری است.")
            seen.add(national_id)
            student = Student(
                organization=job.organization,
                national_id=national_id,
                first_name=str(row["first_name"] or "").strip(),
                last_name=str(row["last_name"] or "").strip(),
                birth_date=_as_date(row["birth_date"], "birth_date"),
                gender=str(row["gender"] or "").strip().lower(),
            )
            student.full_clean()
            prepared.append(student)
        except (ValueError, DjangoValidationError) as exc:
            message = _model_errors(exc) if isinstance(exc, DjangoValidationError) else str(exc)
            errors.append({"row": number, "message": message})
    return prepared, errors


def _validate_enrollments(job, rows):
    prepared, errors, seen = [], [], set()
    normalized_rows = []
    national_ids, year_codes, grade_codes, class_codes = set(), set(), set(), set()
    for number, row in enumerate(rows, start=2):
        national_id = str(row["national_id"] or "").strip().zfill(10)
        year_code = str(row["academic_year_code"] or "").strip()
        grade_code = str(row["grade_code"] or "").strip()
        class_code = str(row["class_code"] or "").strip()
        normalized_rows.append((number, row, national_id, year_code, grade_code, class_code))
        national_ids.add(national_id)
        year_codes.add(year_code)
        grade_codes.add(grade_code)
        class_codes.add(class_code)

    students = {
        item.national_id: item
        for item in Student.objects.filter(
            organization=job.organization, national_id__in=national_ids
        )
    }
    years = {
        item.code: item
        for item in AcademicYear.objects.filter(organization=job.organization, code__in=year_codes)
    }
    grades = {
        item.code: item
        for item in GradeLevel.objects.filter(organization=job.organization, code__in=grade_codes)
    }
    classes = {
        (str(item.academic_year_id), str(item.grade_level_id), item.code): item
        for item in ClassSection.objects.filter(
            school=job.school,
            academic_year__in=years.values(),
            grade_level__in=grades.values(),
            code__in=class_codes,
        )
    }
    active_keys = {
        (str(student_id), str(year_id))
        for student_id, year_id in Enrollment.objects.filter(
            student__in=students.values(),
            academic_year__in=years.values(),
            status=Enrollment.Status.ACTIVE,
        ).values_list("student_id", "academic_year_id")
    }

    for number, row, national_id, year_code, grade_code, class_code in normalized_rows:
        try:
            student = students.get(national_id)
            year = years.get(year_code)
            grade = grades.get(grade_code)
            if student is None or year is None or grade is None:
                raise ValueError("دانش‌آموز، سال، پایه یا کلاس پیدا نشد.")
            class_section = classes.get((str(year.id), str(grade.id), class_code))
            if class_section is None:
                raise ValueError("دانش‌آموز، سال، پایه یا کلاس پیدا نشد.")
            key = (str(student.id), str(year.id))
            if key in seen or key in active_keys:
                raise ValueError("دانش‌آموز در این سال تحصیلی ثبت‌نام فعال دارد.")
            seen.add(key)
            enrollment = Enrollment(
                student=student,
                school=job.school,
                academic_year=year,
                grade_level=grade,
                class_section=class_section,
                student_number=str(row["student_number"] or "").strip(),
                enrolled_on=_as_date(row["enrolled_on"], "enrolled_on"),
            )
            enrollment.full_clean()
            prepared.append(enrollment)
        except (ValueError, DjangoValidationError) as exc:
            message = _model_errors(exc) if isinstance(exc, DjangoValidationError) else str(exc)
            errors.append({"row": number, "message": message})

    class_counts = defaultdict(int)
    for enrollment in prepared:
        class_counts[enrollment.class_section_id] += 1
    current_counts = {
        row["class_section_id"]: row["count"]
        for row in Enrollment.objects.filter(
            class_section_id__in=class_counts,
            status=Enrollment.Status.ACTIVE,
        )
        .values("class_section_id")
        .annotate(count=Count("id"))
    }
    sections = {item.id: item for item in ClassSection.objects.filter(id__in=class_counts)}
    for class_id, added in class_counts.items():
        section = sections[class_id]
        if current_counts.get(class_id, 0) + added > section.capacity:
            errors.append({"row": None, "message": f"ظرفیت کلاس {section.title} کافی نیست."})
    return prepared, errors


def _validate_scores(job, rows):
    grouped, errors, seen = defaultdict(list), [], set()
    assessment_ids = {str(row["assessment_id"]) for row in rows if row.get("assessment_id")}
    assessments = {
        str(item.id): item
        for item in Assessment.objects.select_related("course_offering__class_section").filter(
            id__in=assessment_ids,
            course_offering__class_section__school=job.school,
        )
    }
    national_ids = {str(row["national_id"] or "").strip().zfill(10) for row in rows}
    class_ids = {item.course_offering.class_section_id for item in assessments.values()}
    enrollments = {
        (item.student.national_id, str(item.class_section_id)): item
        for item in Enrollment.objects.select_related("student").filter(
            student__organization=job.organization,
            student__national_id__in=national_ids,
            class_section_id__in=class_ids,
            status=Enrollment.Status.ACTIVE,
        )
    }

    for number, row in enumerate(rows, start=2):
        try:
            assessment = assessments.get(str(row["assessment_id"]))
            if assessment is None:
                raise ValueError("ارزیابی یا ثبت‌نام متناظر پیدا نشد.")
            if assessment.status not in [Assessment.Status.DRAFT, Assessment.Status.REJECTED]:
                raise ValueError("ارزیابی قابل ویرایش نیست.")
            national_id = str(row["national_id"] or "").strip().zfill(10)
            enrollment = enrollments.get(
                (national_id, str(assessment.course_offering.class_section_id))
            )
            if enrollment is None:
                raise ValueError("ارزیابی یا ثبت‌نام متناظر پیدا نشد.")
            key = (assessment.id, enrollment.id)
            if key in seen:
                raise ValueError("نمره تکراری در فایل وجود دارد.")
            seen.add(key)
            status = str(row["status"] or "").strip().lower()
            if status not in Score.Status.values:
                raise ValueError(f"status باید یکی از {', '.join(Score.Status.values)} باشد.")
            value = _as_decimal(row["value"])
            probe = Score(
                assessment=assessment,
                enrollment=enrollment,
                value=value,
                status=status,
                note=str(row["note"] or "").strip(),
                recorded_by=job.requested_by,
            )
            probe.full_clean(exclude=["id"], validate_unique=False)
            grouped[assessment].append(
                {
                    "enrollment": enrollment,
                    "value": value,
                    "status": status,
                    "note": probe.note,
                }
            )
        except (ValueError, DjangoValidationError) as exc:
            message = _model_errors(exc) if isinstance(exc, DjangoValidationError) else str(exc)
            errors.append({"row": number, "message": message})
    return grouped, errors


def process_import_job(job_id):
    with transaction.atomic():
        job = ImportJob.objects.select_for_update().get(pk=job_id)
        stale_before = timezone.now() - timedelta(
            minutes=getattr(settings, "IMPORT_PROCESSING_TIMEOUT_MINUTES", 30)
        )
        if job.status == ImportJob.Status.PROCESSING:
            if job.started_at and job.started_at >= stale_before:
                return job
        elif job.status not in [ImportJob.Status.QUEUED, ImportJob.Status.FAILED]:
            return job
        job.status = ImportJob.Status.PROCESSING
        job.started_at = timezone.now()
        job.finished_at = None
        job.successful_rows = 0
        job.error_count = 0
        job.errors = []
        job.save(
            update_fields=[
                "status",
                "started_at",
                "finished_at",
                "successful_rows",
                "error_count",
                "errors",
                "updated_at",
            ]
        )
    try:
        rows = _load_rows(job)
        validators = {
            ImportJob.ImportType.STUDENTS: _validate_students,
            ImportJob.ImportType.ENROLLMENTS: _validate_enrollments,
            ImportJob.ImportType.SCORES: _validate_scores,
        }
        prepared, errors = validators[job.import_type](job, rows)
        job.total_rows = len(rows)
        if errors:
            job.status = ImportJob.Status.FAILED
            job.error_count = len(errors)
            job.errors = errors[:1000]
            job.finished_at = timezone.now()
            job.save()
            return job
        with transaction.atomic():
            if job.import_type in [ImportJob.ImportType.STUDENTS, ImportJob.ImportType.ENROLLMENTS]:
                if job.import_type == ImportJob.ImportType.ENROLLMENTS:
                    class_ids = sorted({item.class_section_id for item in prepared}, key=str)
                    locked_classes = {
                        item.id: item
                        for item in ClassSection.objects.select_for_update()
                        .filter(id__in=class_ids)
                        .order_by("id")
                    }
                    added_by_class = defaultdict(int)
                    for item in prepared:
                        added_by_class[item.class_section_id] += 1
                    for class_id, added in added_by_class.items():
                        section = locked_classes[class_id]
                        current = Enrollment.objects.filter(
                            class_section=section,
                            status=Enrollment.Status.ACTIVE,
                        ).count()
                        if current + added > section.capacity:
                            raise ValueError(f"ظرفیت کلاس {section.title} کافی نیست.")
                for instance in prepared:
                    instance.save()
                successful = len(prepared)
            else:
                successful = 0
                for assessment, entries in prepared.items():
                    successful += len(
                        bulk_upsert_scores(
                            assessment=assessment, entries=entries, actor=job.requested_by
                        )
                    )
        job.status = ImportJob.Status.COMPLETED
        job.successful_rows = successful
        job.error_count = 0
        job.finished_at = timezone.now()
        job.save()
    except OSError:
        job.status = ImportJob.Status.QUEUED
        job.save(update_fields=["status", "updated_at"])
        raise
    except Exception as exc:
        if isinstance(exc, OSError):
            raise
        with transaction.atomic():
            locked = ImportJob.objects.select_for_update().get(pk=job_id)
            locked.status = ImportJob.Status.FAILED
            locked.error_count = 1
            locked.errors = [{"row": None, "message": str(exc)[:2000]}]
            locked.finished_at = timezone.now()
            locked.save()
            return locked


def mark_import_job_failed(job_id, exc):
    with transaction.atomic():
        job = ImportJob.objects.select_for_update().get(pk=job_id)
        if job.status == ImportJob.Status.COMPLETED:
            return job
        job.status = ImportJob.Status.FAILED
        job.error_count = 1
        job.errors = [{"row": None, "message": str(exc)[:2000]}]
        job.finished_at = timezone.now()
        job.save()
    return job
