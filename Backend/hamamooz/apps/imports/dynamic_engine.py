"""
Dynamic import helpers for Hamamooz workbook templates.

This module avoids fixed column indexes and discovers workbook structure from
semantic headers, periods and indicator codes.
"""

from dataclasses import dataclass, field
from typing import Any


HEADER_ALIASES = {
    "national_id": {"کد ملی", "national_id", "national code"},
    "first_name": {"نام", "first_name"},
    "last_name": {"نام خانوادگی", "نام خانوادگي", "last_name"},
    "class_name": {"کلاس", "نام کلاس", "class"},
    "grade": {"پایه", "grade"},
    "period": {"ماه", "دوره", "نوبت", "period"},
}


@dataclass
class WorkbookProfile:
    sheets: list[str] = field(default_factory=list)
    detected_entities: dict[str, str] = field(default_factory=dict)
    indicators: list[str] = field(default_factory=list)
    periods: list[str] = field(default_factory=list)


class DynamicWorkbookAnalyzer:
    """Analyze XLSX files without assuming one school template version."""

    def normalize(self, value: Any) -> str:
        return str(value or "").strip().lower()

    def detect_headers(self, row):
        result = {}
        normalized_aliases = {
            key: {self.normalize(v) for v in values}
            for key, values in HEADER_ALIASES.items()
        }
        for index, value in enumerate(row):
            text = self.normalize(value)
            for key, aliases in normalized_aliases.items():
                if text in aliases:
                    result[key] = index
        return result

    def analyze(self, workbook):
        profile = WorkbookProfile(sheets=list(workbook.sheetnames))
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 30), values_only=True):
                headers = self.detect_headers(row)
                if headers:
                    profile.detected_entities[sheet.title] = ",".join(headers.keys())
                    break
            for row in sheet.iter_rows(values_only=True):
                for value in row:
                    code = str(value or "")
                    if (code.startswith("EDU_") or code.startswith("PER_")) and code not in profile.indicators:
                        profile.indicators.append(code)
        return profile


def inspect_uploaded_workbook(path):
    from openpyxl import load_workbook
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return DynamicWorkbookAnalyzer().analyze(workbook)
    finally:
        workbook.close()
