from dataclasses import dataclass, field

from openpyxl import load_workbook


@dataclass
class WorkbookSchema:
    file_name: str = ""
    sheets: list[str] = field(default_factory=list)
    columns: dict[str, list[str]] = field(default_factory=dict)
    indicators: list[str] = field(default_factory=list)
    periods: list[str] = field(default_factory=list)
    students: list[dict] = field(default_factory=list)


def _clean(value):
    return str(value or "").strip()


def _is_header(row):
    text = " ".join(_clean(v) for v in row)
    return any(x in text for x in ["کد ملی", "نام خانوادگی", "EDU_", "ماه"])


def analyze_workbook(file_object):
    workbook = load_workbook(file_object, read_only=True, data_only=True)
    try:
        schema = WorkbookSchema(file_name=getattr(file_object, "name", ""))
        schema.sheets = list(workbook.sheetnames)

        for sheet_name in schema.sheets:
            sheet = workbook[sheet_name]
            header = []
            data_start = 0
            for index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                values = [_clean(v) for v in row]
                if _is_header(values):
                    header = values
                    data_start = index
                    break

            schema.columns[sheet_name] = header

            for row in sheet.iter_rows(min_row=data_start + 1, values_only=True):
                for value in row:
                    value = _clean(value)
                    if "|" in value:
                        code = value.split("|", 1)[0].strip()
                        if "_" in code and code not in schema.indicators:
                            schema.indicators.append(code)

            if "ماه" in header:
                idx = header.index("ماه")
                for row in sheet.iter_rows(min_row=data_start + 1, values_only=True):
                    if idx < len(row) and _clean(row[idx]) and _clean(row[idx]) not in schema.periods:
                        schema.periods.append(_clean(row[idx]))

            if "کد ملی" in header:
                idx = header.index("کد ملی")
                for row in sheet.iter_rows(min_row=data_start + 1, values_only=True):
                    if idx < len(row) and _clean(row[idx]):
                        schema.students.append({"national_code": _clean(row[idx])})

        return schema
    finally:
        workbook.close()
