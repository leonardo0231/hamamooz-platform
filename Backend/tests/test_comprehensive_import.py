from datetime import date
from io import BytesIO
from pathlib import Path

import pytest
from django.conf import settings
from django.core.files.base import ContentFile
from django.core.files.uploadedfile import SimpleUploadedFile
from openpyxl import load_workbook

from hamamooz.apps.evaluations.catalog import METRIC_CATALOG
from hamamooz.apps.evaluations.models import MetricScore, MonthlyEvaluation
from hamamooz.apps.imports.models import ImportJob
from hamamooz.apps.imports.pipeline import process_import_job as process_hardened_import_job
from hamamooz.apps.imports.serializers import uploaded_file_checksum
from hamamooz.apps.imports.services import process_import_job
from hamamooz.apps.organizations.models import ClassSection
from hamamooz.apps.students.models import Enrollment, Student


def comprehensive_workbook_bytes(
    base_data,
    *,
    class_code="7-c",
    class_title="هفتم ج",
    capacity=30,
    national_id="0012345680",
    student_number="103",
    score=4,
    metric_count=None,
    academic_year_code=None,
):
    if metric_count is None:
        metric_count = len(METRIC_CATALOG)
    template = (
        Path(settings.BASE_DIR) / "docs" / "import_templates" / "comprehensive_school_template.xlsx"
    )
    workbook = load_workbook(template)
    classes = workbook["کلاس‌بندی"]
    students = workbook["دانش‌آموزان"]
    evaluations = workbook["ثبت اطلاعات"]

    classes["B5"] = base_data["school1"].code
    classes["C5"] = academic_year_code or base_data["year"].code
    classes["D5"] = class_code
    classes["E5"] = class_title
    classes["F5"] = base_data["grade"].title
    classes["G5"] = capacity

    students["C5"] = national_id
    students["D5"] = student_number
    students["E5"] = "سارا"
    students["F5"] = "احمدی"
    students["G5"] = "دختر"
    students["H5"] = date(2012, 3, 4)
    students["I5"] = class_code

    for index in range(metric_count):
        evaluations.cell(row=5, column=7 + index, value=score)
    evaluations.cell(row=5, column=94, value="ثبت از فایل جامع")

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def large_workbook_with_invalid_class_references(base_data):
    """Create the shape that previously produced hundreds of cascade errors."""
    template = (
        Path(settings.BASE_DIR) / "docs" / "import_templates" / "comprehensive_school_template.xlsx"
    )
    workbook = load_workbook(template)
    classes = workbook["کلاس‌بندی"]
    students = workbook["دانش‌آموزان"]
    evaluations = workbook["ثبت اطلاعات"]

    class_codes = []
    for index in range(4):
        row = 5 + index
        class_code = f"invalid-class-{index + 1}"
        class_codes.append(class_code)
        classes.cell(row=row, column=2, value="wrong-school")
        classes.cell(row=row, column=3, value="unknown-year")
        classes.cell(row=row, column=4, value=class_code)
        classes.cell(row=row, column=5, value=f"Invalid class {index + 1}")
        classes.cell(row=row, column=6, value=base_data["grade"].code)
        classes.cell(row=row, column=7, value=30)

    for index in range(100):
        row = 5 + index
        local_code = str(index + 1)
        class_code = class_codes[index % len(class_codes)]
        students.cell(row=row, column=2, value=local_code)
        students.cell(
            row=row,
            column=3,
            value="invalid-national-id" if index == 0 else f"{9_000_000_000 + index:010d}",
        )
        students.cell(row=row, column=4, value=f"bulk-{index + 1:03d}")
        students.cell(row=row, column=5, value="Sara")
        students.cell(row=row, column=6, value="Ahmadi")
        students.cell(row=row, column=7, value="female")
        students.cell(row=row, column=8, value=date(2012, 3, 4))
        students.cell(row=row, column=9, value=class_code)

        evaluations.cell(row=row, column=1, value=index + 1)
        evaluations.cell(row=row, column=2, value=1)
        evaluations.cell(row=row, column=3, value=local_code)
        evaluations.cell(row=row, column=7, value=4)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def create_comprehensive_job(base_data, payload):
    source = ContentFile(payload, name="comprehensive_school.xlsx")
    checksum = uploaded_file_checksum(source)
    return ImportJob.objects.create(
        organization=base_data["organization"],
        school=base_data["school1"],
        import_type=ImportJob.ImportType.COMPREHENSIVE_SCHOOL,
        source_file=source,
        checksum=checksum,
        requested_by=base_data["manager"],
    )


@pytest.mark.django_db
def test_comprehensive_import_creates_class_student_enrollment_and_all_metrics(base_data):
    payload = comprehensive_workbook_bytes(base_data)
    job = create_comprehensive_job(base_data, payload)

    process_import_job(job.id)
    job.refresh_from_db()

    assert job.status == ImportJob.Status.COMPLETED
    assert job.total_rows == 3
    assert job.successful_rows == 3
    section = ClassSection.objects.get(
        school=base_data["school1"], academic_year=base_data["year"], code="7-c"
    )
    student = Student.objects.get(organization=base_data["organization"], national_id="0012345680")
    enrollment = Enrollment.objects.get(student=student, academic_year=base_data["year"])
    assert enrollment.class_section == section
    assert enrollment.student_number == "103"
    evaluation = MonthlyEvaluation.objects.get(enrollment=enrollment, month_no=1)
    assert evaluation.note == "ثبت از فایل جامع"
    assert MetricScore.objects.filter(evaluation=evaluation).count() == len(METRIC_CATALOG)
    assert job.result_summary["classes_created"] == 1
    assert job.result_summary["students_created"] == 1
    assert job.result_summary["enrollments_created"] == 1
    assert job.result_summary["final_evaluations"] == 1
    assert job.result_summary["metric_scores_upserted"] == len(METRIC_CATALOG)


@pytest.mark.django_db
def test_comprehensive_import_marks_partial_evaluation_as_provisional(base_data):
    payload = comprehensive_workbook_bytes(base_data, metric_count=2)
    job = create_comprehensive_job(base_data, payload)

    process_import_job(job.id)
    job.refresh_from_db()

    assert job.status == ImportJob.Status.COMPLETED
    assert job.result_summary["provisional_evaluations"] == 1
    assert job.result_summary["final_evaluations"] == 0
    evaluation = MonthlyEvaluation.objects.get(source_import_job=job)
    assert evaluation.metric_scores.count() == 2


@pytest.mark.django_db
def test_comprehensive_import_validation_error_rolls_back_every_section(base_data):
    payload = comprehensive_workbook_bytes(base_data, score=6)
    job = create_comprehensive_job(base_data, payload)

    process_import_job(job.id)
    job.refresh_from_db()

    assert job.status == ImportJob.Status.FAILED
    assert job.error_count == 1
    assert job.errors[0]["sheet"] == "ثبت اطلاعات"
    assert job.errors[0]["row"] == 5
    assert not ClassSection.objects.filter(school=base_data["school1"], code="7-c").exists()
    assert not Student.objects.filter(national_id="0012345680").exists()
    assert not MonthlyEvaluation.objects.filter(source_import_job=job).exists()


@pytest.mark.django_db
def test_comprehensive_import_runtime_capacity_failure_is_atomic(base_data):
    original_capacity = base_data["class1"].capacity
    payload = comprehensive_workbook_bytes(
        base_data,
        class_code=base_data["class1"].code,
        class_title=base_data["class1"].title,
        capacity=2,
    )
    job = create_comprehensive_job(base_data, payload)

    process_import_job(job.id)
    job.refresh_from_db()
    base_data["class1"].refresh_from_db()

    assert job.status == ImportJob.Status.FAILED
    assert "ظرفیت کلاس" in job.errors[0]["message"]
    assert base_data["class1"].capacity == original_capacity
    assert not Student.objects.filter(national_id="0012345680").exists()
    assert not Enrollment.objects.filter(student_number="103").exists()


@pytest.mark.django_db
def test_comprehensive_import_upserts_existing_records_without_duplicates(base_data):
    first = create_comprehensive_job(base_data, comprehensive_workbook_bytes(base_data, score=3))
    process_import_job(first.id)
    second_payload = comprehensive_workbook_bytes(
        base_data,
        class_title="هفتم ج - به‌روزشده",
        score=5,
    )
    second = create_comprehensive_job(base_data, second_payload)

    process_import_job(second.id)
    second.refresh_from_db()

    assert second.status == ImportJob.Status.COMPLETED
    assert ClassSection.objects.filter(school=base_data["school1"], code="7-c").count() == 1
    assert Student.objects.filter(national_id="0012345680").count() == 1
    assert Enrollment.objects.filter(student__national_id="0012345680").count() == 1
    evaluation = MonthlyEvaluation.objects.get(enrollment__student__national_id="0012345680")
    assert evaluation.metric_scores.count() == len(METRIC_CATALOG)
    assert set(evaluation.metric_scores.values_list("value", flat=True)) == {5}
    assert second.result_summary["classes_updated"] == 1
    assert second.result_summary["students_updated"] == 1
    assert second.result_summary["enrollments_updated"] == 1
    assert second.result_summary["evaluations_updated"] == 1


@pytest.mark.django_db
def test_comprehensive_import_reports_unknown_academic_year_with_location(base_data):
    payload = comprehensive_workbook_bytes(base_data, academic_year_code="unknown-year")
    job = create_comprehensive_job(base_data, payload)

    process_import_job(job.id)
    job.refresh_from_db()

    assert job.status == ImportJob.Status.FAILED
    assert job.errors[0]["sheet"] == "کلاس‌بندی"
    assert job.errors[0]["column"] == "سال تحصیلی"
    assert job.errors[0]["code"] == "academic_year"
    assert not Student.objects.filter(national_id="0012345680").exists()


@pytest.mark.django_db
def test_comprehensive_import_reports_independent_errors_after_class_preflight(base_data):
    payload = large_workbook_with_invalid_class_references(base_data)
    job = create_comprehensive_job(base_data, payload)

    process_hardened_import_job(job.id)
    job.refresh_from_db()

    assert job.status == ImportJob.Status.FAILED
    assert job.total_rows == 204
    assert job.error_count == 2, job.errors
    assert "structural_validation" not in [error["code"] for error in job.errors]
    assert any(error["sheet"] == "دانش‌آموزان" for error in job.errors)
    assert any(error["sheet"] == "ثبت اطلاعات" for error in job.errors)
    warning_codes = {warning["code"] for warning in job.result_summary["normalization_warnings"]}
    assert "school_code_ignored" in warning_codes
    assert "academic_year_fallback" in warning_codes
    assert not Student.objects.filter(national_id__startswith="9000000").exists()


@pytest.mark.django_db
def test_comprehensive_import_rejects_non_empty_rows_beyond_template_limits(base_data):
    workbook = load_workbook(BytesIO(comprehensive_workbook_bytes(base_data)))
    classes = workbook["کلاس‌بندی"]
    classes.cell(row=35, column=2, value=base_data["school1"].code)
    classes.cell(row=35, column=3, value=base_data["year"].code)
    classes.cell(row=35, column=4, value="overflow-class")
    classes.cell(row=35, column=5, value="Overflow class")
    classes.cell(row=35, column=6, value=base_data["grade"].code)
    classes.cell(row=35, column=7, value=30)
    payload = BytesIO()
    workbook.save(payload)

    job = create_comprehensive_job(base_data, payload.getvalue())
    process_hardened_import_job(job.id)
    job.refresh_from_db()

    assert job.status == ImportJob.Status.FAILED
    assert job.error_count == 1
    assert not ClassSection.objects.filter(code="overflow-class").exists()


@pytest.mark.django_db
def test_comprehensive_template_and_upload_api_contract(api_client, base_data):
    api_client.force_authenticate(base_data["manager"])
    scope = {"HTTP_X_SCHOOL_ID": str(base_data["school1"].id)}

    template_response = api_client.get("/api/v1/imports/templates/comprehensive_school/", **scope)
    template_payload = b"".join(template_response.streaming_content)
    assert template_response.status_code == 200
    workbook = load_workbook(BytesIO(template_payload), read_only=True)
    assert {"کلاس‌بندی", "دانش‌آموزان", "ثبت اطلاعات"}.issubset(workbook.sheetnames)

    payload = comprehensive_workbook_bytes(base_data, metric_count=2)
    response = api_client.post(
        "/api/v1/imports/",
        {
            "school": str(base_data["school1"].id),
            "import_type": ImportJob.ImportType.COMPREHENSIVE_SCHOOL,
            "source_file": SimpleUploadedFile(
                "comprehensive_school.xlsx",
                payload,
                content_type=("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
            ),
        },
        format="multipart",
        **scope,
    )
    assert response.status_code == 201
    job = ImportJob.objects.get(pk=response.data["id"])
    assert response.data["result_summary"] == {}
    process_hardened_import_job(job.id)
    detail = api_client.get(f"/api/v1/imports/{job.id}/", **scope)
    assert detail.status_code == 200
    assert detail.data["result_summary"]["provisional_evaluations"] == 1


@pytest.mark.django_db
def test_comprehensive_upload_rejects_legacy_xls_extension(api_client, base_data):
    api_client.force_authenticate(base_data["manager"])
    response = api_client.post(
        "/api/v1/imports/",
        {
            "school": str(base_data["school1"].id),
            "import_type": ImportJob.ImportType.COMPREHENSIVE_SCHOOL,
            "source_file": SimpleUploadedFile(
                "comprehensive_school.xls",
                b"legacy",
                content_type="application/vnd.ms-excel",
            ),
        },
        format="multipart",
        HTTP_X_SCHOOL_ID=str(base_data["school1"].id),
    )
    assert response.status_code == 400
    assert "XLSX" in str(response.data)


@pytest.mark.django_db
def test_comprehensive_import_cannot_move_active_enrollment_from_another_school(base_data):
    other_student = Student.objects.create(
        organization=base_data["organization"],
        national_id="0012345680",
        first_name="مدرسه",
        last_name="دیگر",
        birth_date=date(2012, 3, 4),
        gender=Student.Gender.FEMALE,
    )
    other_enrollment = Enrollment.objects.create(
        student=other_student,
        school=base_data["school2"],
        academic_year=base_data["year"],
        grade_level=base_data["grade"],
        class_section=base_data["class2"],
        student_number="other-103",
        enrolled_on=base_data["year"].starts_on,
    )
    job = create_comprehensive_job(base_data, comprehensive_workbook_bytes(base_data))

    process_hardened_import_job(job.id)
    job.refresh_from_db()
    other_student.refresh_from_db()
    other_enrollment.refresh_from_db()

    assert job.status == ImportJob.Status.FAILED
    assert "مدرسه دیگری" in job.errors[0]["message"]
    assert other_student.first_name == "مدرسه"
    assert other_enrollment.school == base_data["school2"]
    assert not ClassSection.objects.filter(school=base_data["school1"], code="7-c").exists()
