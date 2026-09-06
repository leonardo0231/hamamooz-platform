from io import BytesIO

import pytest
from django.core.files.base import ContentFile
from openpyxl import load_workbook

from hamamooz.apps.accounts.models import User
from hamamooz.apps.evaluations.models import MonthlyEvaluation
from hamamooz.apps.imports.comprehensive_template import build_comprehensive_school_template
from hamamooz.apps.imports.models import ImportJob
from hamamooz.apps.imports.pipeline import process_import_job
from hamamooz.apps.imports.serializers import uploaded_file_checksum
from hamamooz.apps.organizations.models import AcademicYear, GradeLevel, Organization, School, Term
from hamamooz.apps.students.models import Student


def _operational_summer_workbook(school_code):
    workbook = load_workbook(build_comprehensive_school_template())
    classes = workbook["کلاس‌بندی"]
    students = workbook["دانش‌آموزان"]
    evaluations = workbook["ثبت اطلاعات"]

    classes["B5"] = school_code
    classes["C5"] = "1405-1406"
    classes["D5"] = "805"
    classes["E5"] = "خوارزمی5"
    classes["F5"] = "هشتم"
    classes["G5"] = 35

    students["C5"] = "0960619062"
    students["D5"] = "805-001"
    students["E5"] = "دانش‌آموز"
    students["F5"] = "معتبر"
    students["G5"] = "پسر"
    students["H5"] = "1392/05/06"
    students["I5"] = "805"

    # Mirrors the uploaded operational workbook: one source-domain value, one
    # delta, normal 0..5 rubric scores and a literal not-applicable marker.
    evaluations["G7"] = 19.78
    evaluations["H7"] = -8.78
    evaluations["I7"] = 5
    evaluations["AD7"] = "ندارد"
    evaluations["CP7"] = "داده تابستانی"

    # One intentionally incomplete identity row must not block the other rows.
    students["C6"] = "93100753121"
    students["D6"] = "805-002"
    students["E6"] = "ردیف"
    students["F6"] = "ناقص"
    students["G6"] = "پسر"
    students["H6"] = "1392/06/01"
    students["I6"] = "805"
    evaluations["I19"] = 5

    recommendations = workbook.create_sheet("بانک راهکارها")
    recommendations.append(["بانک راهکارها"])
    recommendations.append([])
    recommendations.append([])
    recommendations.append(["ردیف", "حوزه", "راهکار", "نحوه استفاده"])
    recommendations.append([1, "آموزشی", "تمرین هدفمند", "پس از بررسی انسانی"])

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


@pytest.mark.django_db
def test_comprehensive_import_bootstraps_summer_and_preserves_operational_values():
    organization = Organization.objects.create(name="مدرسه آزمایشی", code="template-import-org")
    school = School.objects.create(
        organization=organization,
        code="95114129",
        name="شعبه تابستانی",
    )
    manager = User.objects.create_user(
        username="template-import-manager",
        email="template-import@example.com",
        password="Strong-pass-123",
    )

    assert not AcademicYear.objects.filter(organization=organization).exists()
    assert not GradeLevel.objects.filter(organization=organization).exists()
    assert not Term.objects.exists()

    source = ContentFile(
        _operational_summer_workbook(school.code),
        name="805.806.807.xlsx",
    )
    checksum = uploaded_file_checksum(source)
    job = ImportJob.objects.create(
        organization=organization,
        school=school,
        import_type=ImportJob.ImportType.COMPREHENSIVE_SCHOOL,
        source_file=source,
        checksum=checksum,
        requested_by=manager,
    )

    process_import_job(job.id)
    job.refresh_from_db()

    assert job.status == ImportJob.Status.COMPLETED, job.errors
    year = AcademicYear.objects.get(organization=organization, code="1405-1406")
    grade = GradeLevel.objects.get(organization=organization, order=8)
    assert grade.title == "هشتم"

    terms = list(Term.objects.filter(academic_year=year).order_by("order"))
    assert [item.code for item in terms] == [Term.Code.SUMMER]
    assert terms[0].title == "تابستان"

    student = Student.objects.get(organization=organization, national_id="0960619062")
    assert not Student.objects.filter(organization=organization, national_id="93100753121").exists()
    evaluation = MonthlyEvaluation.objects.get(enrollment__student=student, month_no=3)
    assert evaluation.term == terms[0]
    assert evaluation.raw_metric_values["EDU_01"] == pytest.approx(19.78)
    assert evaluation.raw_metric_values["EDU_02"] == pytest.approx(-8.78)
    assert evaluation.raw_metric_values["EDU_03"] == 5
    assert evaluation.raw_metric_values["RES_01"] == "ندارد"
    assert evaluation.metric_scores.get(metric_code="EDU_03").value == 5
    assert not evaluation.metric_scores.filter(metric_code__in=["EDU_01", "EDU_02", "RES_01"]).exists()

    reference = job.result_summary["reference_data"]
    assert reference["academic_year_created"] == 1
    assert reference["grades_created"] == 1
    assert reference["detected_term_codes"] == [Term.Code.SUMMER]
    assert job.result_summary["quarantined_student_count"] == 1
    assert len(job.result_summary["workbook_profile"]["domain_weights"]) == 9
    assert job.result_summary["workbook_profile"]["recommendations"][0]["domain"] == "آموزشی"

    warning_codes = {
        item["code"] for item in job.result_summary.get("normalization_warnings", [])
    }
    assert {
        "metric_value_preserved_raw",
        "metric_not_applicable",
        "student_row_quarantined",
        "evaluation_row_quarantined",
    }.issubset(warning_codes)
