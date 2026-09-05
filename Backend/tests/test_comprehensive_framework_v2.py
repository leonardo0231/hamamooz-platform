from openpyxl import load_workbook

from hamamooz.apps.evaluations.catalog import FRAMEWORK_VERSION, METRIC_CATALOG
from hamamooz.apps.imports.comprehensive import _metric_columns
from hamamooz.apps.imports.comprehensive_template import build_comprehensive_school_template


def test_current_template_encodes_46_metric_layout():
    workbook = load_workbook(build_comprehensive_school_template(), read_only=True, data_only=False)
    sheet = workbook["ثبت اطلاعات"]

    metric_columns, framework_version = _metric_columns(sheet)

    assert framework_version == FRAMEWORK_VERSION == "2.0"
    assert len(METRIC_CATALOG) == len(metric_columns) == 46
    assert metric_columns[7] == "EDU_01"
    assert metric_columns[14] == "EDU_08"
    assert metric_columns[15] == "DEV_01"
    assert metric_columns[52] == "PER_03"
    assert all(sheet.cell(4, column).value is None for column in range(53, 81))
    assert sheet["CC4"].value == "امتیاز آموزشی (۰ تا ۲۰)"
    assert sheet["CP4"].value == "توضیحات"
    workbook.close()
