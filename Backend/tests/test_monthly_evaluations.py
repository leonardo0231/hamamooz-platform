from io import BytesIO

import pytest
from django.core.files.base import ContentFile
from openpyxl import Workbook, load_workbook
from test_imports import create_job

from hamamooz.apps.evaluations.catalog import DOMAIN_DEFINITIONS, METRIC_CATALOG
from hamamooz.apps.evaluations.models import MetricScore, MonthlyEvaluation
from hamamooz.apps.imports.models import ImportJob
from hamamooz.apps.imports.serializers import uploaded_file_checksum
from hamamooz.apps.imports.services import process_import_job
from hamamooz.apps.imports.templates import build_smart_evaluation_template


def evaluation_row(base_data, metric_code, score, *, month_no=4, school_code=None):
    enrollment = base_data["enrollments"][0]
    return [
        "1.0",
        school_code or base_data["school1"].code,
        base_data["year"].code,
        base_data["class1"].code,
        enrollment.student_number,
        enrollment.student.national_id,
        month_no,
        metric_code,
        score,
        "ارزیابی مهر",
    ]


def smart_workbook_bytes(base_data, metric_values, *, school_code=None):
    workbook = Workbook()
    guide = workbook.active
    guide.title = "راهنما"
    guide.append(["قالب هوشمند هم‌آموز"])
    workbook.create_sheet("تنظیمات وزن‌دهی")
    students = workbook.create_sheet("دانش آموزان")
    students.append(["ردیف", "کد دانش‌آموزی", "نام و نام خانوادگی", "کلاس"])
    enrollment = base_data["enrollments"][0]
    students.append([1, "1", enrollment.student.full_name, enrollment.class_section.title])

    data = workbook.create_sheet("ثبت اطلاعات")
    data.append([None] * 92)
    metric_codes = list(METRIC_CATALOG)
    data.append(
        ["ردیف", "ماه", "کد دانش‌آموزی", "نام و نام خانوادگی", "کلاس"]
        + [METRIC_CATALOG[code]["title"] for code in metric_codes]
        + [f"میانگین {title} (۲۰)" for title, _weight in DOMAIN_DEFINITIONS.values()]
        + ["نمره وزنی نهایی (۲۰)", "سطح عملکرد", "شماره ماه", "توضیحات"]
    )
    data.append(
        [1, "مهر", "1", enrollment.student.full_name, enrollment.class_section.title]
        + [metric_values.get(code) for code in metric_codes]
        + [19] * 9
        + [19, "عالی", 4, "ثبت هوشمند"]
    )

    metadata = workbook.create_sheet("__hamamooz_meta")
    for row in [
        ["template_version", "2.0"],
        ["framework_version", "1.0"],
        ["school_code", school_code or base_data["school1"].code],
        ["academic_year_code", base_data["year"].code],
        ["class_code", base_data["class1"].code],
    ]:
        metadata.append(row)
    metadata["D1"] = "local_student_code"
    metadata["E1"] = "enrollment_id"
    metadata["F1"] = "national_id"
    metadata["G1"] = "student_number"
    metadata.append([])
    metadata["D2"] = "1"
    metadata["E2"] = str(enrollment.id)
    metadata["F2"] = enrollment.student.national_id
    metadata["G2"] = enrollment.student_number
    metadata["I1"] = "excel_column"
    metadata["J1"] = "metric_code"
    metadata["K1"] = "metric_title"
    for index, code in enumerate(metric_codes, start=2):
        metadata.cell(index, 9, index + 4)
        metadata.cell(index, 10, code)
        metadata.cell(index, 11, METRIC_CATALOG[code]["title"])
    metadata.sheet_state = "veryHidden"

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def create_smart_job(base_data, metric_values, *, school_code=None):
    source = ContentFile(
        smart_workbook_bytes(base_data, metric_values, school_code=school_code),
        name="smart-monthly-evaluations.xlsx",
    )
    checksum = uploaded_file_checksum(source)
    return ImportJob.objects.create(
        organization=base_data["organization"],
        school=base_data["school1"],
        import_type=ImportJob.ImportType.MONTHLY_EVALUATIONS,
        source_file=source,
        checksum=checksum,
        requested_by=base_data["manager"],
    )


def create_job_from_payload(base_data, payload):
    source = ContentFile(payload, name="generated-smart-template.xlsx")
    checksum = uploaded_file_checksum(source)
    return ImportJob.objects.create(
        organization=base_data["organization"],
        school=base_data["school1"],
        import_type=ImportJob.ImportType.MONTHLY_EVALUATIONS,
        source_file=source,
        checksum=checksum,
        requested_by=base_data["manager"],
    )


@pytest.mark.django_db
def test_monthly_evaluation_import_persists_and_updates_scores(base_data):
    job = create_job(
        base_data,
        [
            evaluation_row(base_data, "EDU_01", 4),
            evaluation_row(base_data, "EDU_02", 3),
            evaluation_row(base_data, "DEV_01", 5),
        ],
        ImportJob.ImportType.MONTHLY_EVALUATIONS,
    )

    process_import_job(job.id)
    job.refresh_from_db()

    assert job.status == ImportJob.Status.COMPLETED, job.errors
    assert job.successful_rows == 3
    evaluation = MonthlyEvaluation.objects.get(enrollment=base_data["enrollments"][0], month_no=4)
    assert evaluation.note == "ارزیابی مهر"
    assert evaluation.source_import_job == job
    assert dict(evaluation.metric_scores.values_list("metric_code", "value")) == {
        "DEV_01": 5,
        "EDU_01": 4,
        "EDU_02": 3,
    }

    update_job = create_job(
        base_data,
        [evaluation_row(base_data, "EDU_01", 2)],
        ImportJob.ImportType.MONTHLY_EVALUATIONS,
    )
    process_import_job(update_job.id)

    assert MonthlyEvaluation.objects.count() == 1
    assert MetricScore.objects.get(evaluation=evaluation, metric_code="EDU_01").value == 2


@pytest.mark.django_db
def test_monthly_evaluation_import_is_atomic_on_invalid_row(base_data):
    job = create_job(
        base_data,
        [
            evaluation_row(base_data, "EDU_01", 4),
            evaluation_row(base_data, "UNKNOWN", 3),
        ],
        ImportJob.ImportType.MONTHLY_EVALUATIONS,
    )

    process_import_job(job.id)
    job.refresh_from_db()

    assert job.status == ImportJob.Status.FAILED
    assert MonthlyEvaluation.objects.count() == 0
    assert job.errors[0]["row"] == 3


@pytest.mark.django_db
def test_monthly_evaluation_import_rejects_wrong_school_scope(base_data):
    job = create_job(
        base_data,
        [
            evaluation_row(
                base_data,
                "EDU_01",
                4,
                school_code=base_data["school2"].code,
            )
        ],
        ImportJob.ImportType.MONTHLY_EVALUATIONS,
    )

    process_import_job(job.id)
    job.refresh_from_db()

    assert job.status == ImportJob.Status.FAILED
    assert "شعبه انتخاب‌شده" in job.errors[0]["message"]
    assert MonthlyEvaluation.objects.count() == 0


@pytest.mark.django_db
def test_monthly_evaluation_api_is_scoped_and_returns_calculations(api_client, base_data):
    job = create_job(
        base_data,
        [
            evaluation_row(base_data, "EDU_01", 4),
            evaluation_row(base_data, "EDU_02", 3),
        ],
        ImportJob.ImportType.MONTHLY_EVALUATIONS,
    )
    process_import_job(job.id)
    api_client.force_authenticate(base_data["manager"])

    response = api_client.get(
        "/api/v1/monthly-evaluations/",
        {"enrollment__student": str(base_data["students"][0].id)},
        HTTP_X_SCHOOL_ID=str(base_data["school1"].id),
    )

    assert response.status_code == 200
    result = response.data["results"][0]
    assert result["student"] == str(base_data["students"][0].id)
    assert result["month_no"] == 4
    assert result["overall_score"] == 14.0
    assert result["completion_percent"] == pytest.approx(2 / 74 * 100, abs=0.01)

    denied = api_client.get(
        "/api/v1/monthly-evaluations/",
        HTTP_X_SCHOOL_ID=str(base_data["school2"].id),
    )
    assert denied.status_code == 403


@pytest.mark.django_db
def test_smart_wide_import_reads_named_sheet_and_unpivots_metrics(base_data):
    job = create_smart_job(base_data, {"EDU_01": 5, "DEV_01": 4})

    process_import_job(job.id)
    job.refresh_from_db()

    assert job.status == ImportJob.Status.COMPLETED
    assert job.total_rows == 1
    assert job.successful_rows == 1
    evaluation = MonthlyEvaluation.objects.get(enrollment=base_data["enrollments"][0], month_no=4)
    assert evaluation.note == "ثبت هوشمند"
    assert dict(evaluation.metric_scores.values_list("metric_code", "value")) == {
        "DEV_01": 4,
        "EDU_01": 5,
    }


@pytest.mark.django_db
def test_smart_wide_import_rejects_tampered_school_metadata(base_data):
    job = create_smart_job(
        base_data,
        {"EDU_01": 5},
        school_code=base_data["school2"].code,
    )

    process_import_job(job.id)
    job.refresh_from_db()

    assert job.status == ImportJob.Status.FAILED
    assert "شعبه انتخاب‌شده" in job.errors[0]["message"]
    assert MonthlyEvaluation.objects.count() == 0


@pytest.mark.django_db
def test_partial_evaluation_is_provisional_and_hides_final_only_analytics(api_client, base_data):
    job = create_job(
        base_data,
        [evaluation_row(base_data, "EDU_01", 5)],
        ImportJob.ImportType.MONTHLY_EVALUATIONS,
    )
    process_import_job(job.id)
    api_client.force_authenticate(base_data["manager"])

    response = api_client.get(
        "/api/v1/monthly-evaluations/",
        HTTP_X_SCHOOL_ID=str(base_data["school1"].id),
    )

    assert response.status_code == 200
    result = response.data["results"][0]
    assert result["completion_status"] == "provisional"
    assert result["performance_level"] is None
    assert result["completion_warning"]


@pytest.mark.django_db
def test_student_analytics_returns_trend_strengths_and_real_rank_count(api_client, base_data):
    rows = []
    for enrollment_index, enrollment in enumerate(base_data["enrollments"]):
        for month_no, score in [(4, 3 + enrollment_index), (5, 5 - enrollment_index)]:
            for metric_code in METRIC_CATALOG:
                rows.append(
                    [
                        "1.0",
                        base_data["school1"].code,
                        base_data["year"].code,
                        base_data["class1"].code,
                        enrollment.student_number,
                        enrollment.student.national_id,
                        month_no,
                        metric_code,
                        score,
                        "",
                    ]
                )
    job = create_job(base_data, rows, ImportJob.ImportType.MONTHLY_EVALUATIONS)
    process_import_job(job.id)
    api_client.force_authenticate(base_data["manager"])

    response = api_client.get(
        "/api/v1/monthly-evaluations/analytics/",
        {"enrollment": str(base_data["enrollments"][0].id), "rank_scope": "school"},
        HTTP_X_SCHOOL_ID=str(base_data["school1"].id),
    )

    assert response.status_code == 200
    assert response.data["completion_status"] == "final"
    assert response.data["first_month"] == 4
    assert response.data["last_month"] == 5
    assert response.data["change"] == 8.0
    assert response.data["trend"] == "improving"
    assert response.data["strongest_domain"] is not None
    assert response.data["weakest_domain"] is not None
    assert response.data["recommendation"]
    assert response.data["rank"] == 1
    assert response.data["ranked_count"] == 2

    dashboard = api_client.get(
        "/api/v1/monthly-evaluations/dashboard/",
        {"academic_year": str(base_data["year"].id)},
        HTTP_X_SCHOOL_ID=str(base_data["school1"].id),
    )
    assert dashboard.status_code == 200
    assert dashboard.data["counts"] == {
        "students": 2,
        "evaluated": 2,
        "final": 2,
        "provisional": 0,
        "ranked": 2,
    }
    assert dashboard.data["monthly_trend"] == [
        {"month_no": 4, "average": 14.0, "students": 2},
        {"month_no": 5, "average": 18.0, "students": 2},
    ]

    export = api_client.get(
        "/api/v1/monthly-evaluations/export/",
        {"academic_year": str(base_data["year"].id)},
        HTTP_X_SCHOOL_ID=str(base_data["school1"].id),
    )
    assert export.status_code == 200
    export_workbook = load_workbook(
        BytesIO(b"".join(export.streaming_content)),
        data_only=True,
    )
    assert export_workbook.sheetnames == ["ثبت اطلاعات", "تحلیل دانش آموزان", "داشبورد"]
    assert export_workbook["ثبت اطلاعات"].max_row == 5
    assert export_workbook["تحلیل دانش آموزان"]["O2"].value == 1
    assert export_workbook["داشبورد"]["B2"].value == 2
    export_workbook.close()


@pytest.mark.django_db
def test_internal_smart_template_builder_embeds_stable_metadata(base_data):
    template = build_smart_evaluation_template(base_data["class1"])
    template.seek(0)
    workbook = load_workbook(template, data_only=False)
    assert workbook["__hamamooz_meta"].sheet_state == "veryHidden"
    assert workbook["__hamamooz_meta"]["B1"].value == "2.0"
    assert workbook["__hamamooz_meta"]["B3"].value == base_data["school1"].code
    assert workbook["ثبت اطلاعات"].max_row == 2 + len(base_data["enrollments"]) * 12
    workbook.close()


@pytest.mark.django_db
def test_internal_generated_smart_template_round_trips_for_historical_jobs(base_data):
    template = build_smart_evaluation_template(base_data["class1"])
    template.seek(0)
    workbook = load_workbook(template)
    workbook["ثبت اطلاعات"]["F3"] = 5
    workbook["ثبت اطلاعات"]["CN3"] = "رفت و برگشت قالب"
    payload = BytesIO()
    workbook.save(payload)
    workbook.close()

    job = create_job_from_payload(base_data, payload.getvalue())
    process_import_job(job.id)
    job.refresh_from_db()

    assert job.status == ImportJob.Status.COMPLETED, job.errors
    evaluation = MonthlyEvaluation.objects.get(month_no=1)
    assert evaluation.note == "رفت و برگشت قالب"
    assert evaluation.metric_scores.get(metric_code="EDU_01").value == 5
