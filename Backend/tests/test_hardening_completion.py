from io import BytesIO
from pathlib import Path

import pytest
from django.conf import settings
from django.core.files.base import ContentFile
from django.core.files.uploadedfile import SimpleUploadedFile
from openpyxl import load_workbook

from hamamooz.apps.core.models import AuditEvent
from hamamooz.apps.evaluations.catalog import METRIC_CATALOG
from hamamooz.apps.evaluations.models import MonthlyEvaluation
from hamamooz.apps.imports.models import ImportJob
from hamamooz.apps.imports.pipeline import process_import_job
from hamamooz.apps.imports.serializers import uploaded_file_checksum
from hamamooz.apps.students.models import Student


def comprehensive_payload(
    base_data,
    *,
    national_id="0012345680",
    evaluation_national_id=None,
    evaluation_name=None,
    evaluation_class_code=None,
    metric_count=2,
    score=4,
    marker=None,
):
    template = (
        Path(settings.BASE_DIR)
        / "docs"
        / "import_templates"
        / "comprehensive_school_template.xlsx"
    )
    workbook = load_workbook(template)
    classes = workbook["کلاس‌بندی"]
    students = workbook["دانش‌آموزان"]
    evaluations = workbook["ثبت اطلاعات"]

    class_code = "7-hardening"
    classes["B5"] = base_data["school1"].code
    classes["C5"] = base_data["year"].code
    classes["D5"] = class_code
    classes["E5"] = "هفتم سخت‌سازی"
    classes["F5"] = base_data["grade"].title
    classes["G5"] = 30

    students["B5"] = "1"
    students["C5"] = national_id
    students["D5"] = "hard-103"
    students["E5"] = "سارا"
    students["F5"] = "احمدی"
    students["G5"] = "دختر"
    students["H5"] = "2012-03-04"
    students["I5"] = class_code

    evaluations["C5"] = "1"
    evaluations["D5"] = evaluation_national_id or national_id
    evaluations["E5"] = evaluation_name or "سارا احمدی"
    evaluations["F5"] = evaluation_class_code or class_code
    for index in range(metric_count):
        evaluations.cell(row=5, column=7 + index, value=score)
    evaluations.cell(row=5, column=94, value="ثبت تست سخت‌سازی")

    if marker is not None:
        workbook.worksheets[0]["Z50"] = marker

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def create_comprehensive_job(base_data, payload):
    source = ContentFile(payload, name="comprehensive_school.xlsx")
    checksum = uploaded_file_checksum(source)
    return ImportJob.objects.create(
        organization=base_data["organization"],
        school=base_data["school1"],
        import_type=ImportJob.ImportType.COMPREHENSIVE_SCHOOL,
        source_file=source,
        checksum=checksum,
        requested_by=base_data["manager"],
    )


def manual_payload(base_data, *, month=4, metrics=None, note="ثبت دستی"):
    if metrics is None:
        metrics = [{"metric_code": "EDU_01", "value": 4}]
    return {
        "enrollment": str(base_data["enrollments"][0].id),
        "month_no": month,
        "note": note,
        "metrics": metrics,
    }


@pytest.mark.django_db
def test_public_import_rejects_all_legacy_import_types(api_client, base_data):
    api_client.force_authenticate(base_data["manager"])
    payload = SimpleUploadedFile(
        "students.xlsx",
        b"not-used-because-type-is-rejected-first",
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    response = api_client.post(
        "/api/v1/imports/",
        {
            "school": str(base_data["school1"].id),
            "import_type": ImportJob.ImportType.STUDENTS,
            "source_file": payload,
        },
        format="multipart",
        HTTP_X_SCHOOL_ID=str(base_data["school1"].id),
    )

    assert response.status_code == 400
    assert "فایل جامع" in str(response.data)
    assert ImportJob.objects.count() == 0


@pytest.mark.django_db
def test_hardened_pipeline_rejects_short_national_id_without_padding(base_data):
    payload = comprehensive_payload(base_data, national_id="123456789")
    job = create_comprehensive_job(base_data, payload)

    process_import_job(job.id)
    job.refresh_from_db()

    assert job.status == ImportJob.Status.FAILED
    assert any(error.get("code") == "national_id_format" for error in job.errors)
    assert not Student.objects.filter(first_name="سارا", last_name="احمدی").exists()


@pytest.mark.django_db
@pytest.mark.parametrize(
    ("override", "value", "expected_code"),
    [
        ("evaluation_national_id", "0012345681", "evaluation_identity_mismatch"),
        ("evaluation_name", "شخص دیگر", "evaluation_name_mismatch"),
        ("evaluation_class_code", "7-other", "evaluation_class_mismatch"),
    ],
)
def test_hardened_pipeline_cross_checks_visible_evaluation_identity(
    base_data, override, value, expected_code
):
    payload = comprehensive_payload(base_data, **{override: value})
    job = create_comprehensive_job(base_data, payload)

    process_import_job(job.id)
    job.refresh_from_db()

    assert job.status == ImportJob.Status.FAILED
    assert any(error.get("code") == expected_code for error in job.errors)
    assert not Student.objects.filter(national_id="0012345680").exists()


@pytest.mark.django_db
def test_comprehensive_result_tracks_unchanged_without_deleting_omitted_records(base_data):
    first = create_comprehensive_job(
        base_data, comprehensive_payload(base_data, marker="first")
    )
    process_import_job(first.id)
    first.refresh_from_db()
    assert first.status == ImportJob.Status.COMPLETED, first.errors

    existing_student = base_data["students"][0]
    second = create_comprehensive_job(
        base_data, comprehensive_payload(base_data, marker="second")
    )
    process_import_job(second.id)
    second.refresh_from_db()
    existing_student.refresh_from_db()

    assert second.status == ImportJob.Status.COMPLETED, second.errors
    assert second.result_summary["classes_unchanged"] == 1
    assert second.result_summary["students_unchanged"] == 1
    assert second.result_summary["enrollments_unchanged"] == 1
    assert second.result_summary["evaluations_unchanged"] == 1
    assert second.result_summary["metric_scores_unchanged"] == 2
    assert second.result_summary["records_deleted"] == 0
    assert second.result_summary["delete_policy"] == "explicit_manual_only"
    assert not existing_student.is_deleted
    assert Student.objects.filter(pk=existing_student.pk).exists()


@pytest.mark.django_db
def test_manual_evaluation_update_preserves_omitted_metrics_and_source(api_client, base_data):
    source = ImportJob.objects.create(
        organization=base_data["organization"],
        school=base_data["school1"],
        import_type=ImportJob.ImportType.COMPREHENSIVE_SCHOOL,
        source_file=ContentFile(b"source", name="source.xlsx"),
        checksum="manual-provenance-source",
        requested_by=base_data["manager"],
        status=ImportJob.Status.COMPLETED,
    )
    evaluation = MonthlyEvaluation.objects.create(
        enrollment=base_data["enrollments"][0],
        month_no=4,
        note="از فایل",
        recorded_by=base_data["manager"],
        source_import_job=source,
    )
    evaluation.metric_scores.create(metric_code="EDU_01", value=2)
    evaluation.metric_scores.create(metric_code="DEV_01", value=5)

    api_client.force_authenticate(base_data["manager"])
    response = api_client.post(
        "/api/v1/monthly-evaluations/manual/",
        manual_payload(
            base_data,
            metrics=[{"metric_code": "EDU_01", "value": 4}],
            note="اصلاح دستی",
        ),
        format="json",
        HTTP_X_SCHOOL_ID=str(base_data["school1"].id),
    )

    assert response.status_code == 200
    assert response.data["result"] == {
        "created": False,
        "restored": False,
        "metrics_created": 0,
        "metrics_updated": 1,
        "metrics_unchanged": 0,
    }
    evaluation.refresh_from_db()
    assert evaluation.source_import_job == source
    assert evaluation.note == "اصلاح دستی"
    assert dict(evaluation.metric_scores.values_list("metric_code", "value")) == {
        "DEV_01": 5,
        "EDU_01": 4,
    }
    assert AuditEvent.objects.filter(
        action="evaluation.manual_upserted", entity_id=str(evaluation.id)
    ).exists()


@pytest.mark.django_db
def test_manual_evaluation_validates_payload_and_duplicate_metrics(api_client, base_data):
    api_client.force_authenticate(base_data["manager"])
    scope = {"HTTP_X_SCHOOL_ID": str(base_data["school1"].id)}

    empty = api_client.post(
        "/api/v1/monthly-evaluations/manual/",
        manual_payload(base_data, metrics=[], note=""),
        format="json",
        **scope,
    )
    duplicate = api_client.post(
        "/api/v1/monthly-evaluations/manual/",
        manual_payload(
            base_data,
            metrics=[
                {"metric_code": "EDU_01", "value": 3},
                {"metric_code": "EDU_01", "value": 4},
            ],
        ),
        format="json",
        **scope,
    )

    assert empty.status_code == 400
    assert duplicate.status_code == 400
    assert MonthlyEvaluation.objects.count() == 0


@pytest.mark.django_db
def test_manual_evaluation_prevents_cross_school_write(api_client, base_data):
    api_client.force_authenticate(base_data["teacher2"])

    response = api_client.post(
        "/api/v1/monthly-evaluations/manual/",
        manual_payload(base_data),
        format="json",
        HTTP_X_SCHOOL_ID=str(base_data["school2"].id),
    )

    assert response.status_code == 400
    assert "حوزه دسترسی" in str(response.data)
    assert MonthlyEvaluation.objects.count() == 0


@pytest.mark.django_db
def test_manual_evaluation_delete_is_audited_and_restorable(api_client, base_data):
    api_client.force_authenticate(base_data["manager"])
    scope = {"HTTP_X_SCHOOL_ID": str(base_data["school1"].id)}
    created = api_client.post(
        "/api/v1/monthly-evaluations/manual/",
        manual_payload(base_data),
        format="json",
        **scope,
    )
    assert created.status_code == 201
    evaluation_id = created.data["evaluation"]["id"]

    missing_reason = api_client.delete(
        f"/api/v1/monthly-evaluations/{evaluation_id}/manual/",
        {},
        format="json",
        **scope,
    )
    assert missing_reason.status_code == 400

    deleted = api_client.delete(
        f"/api/v1/monthly-evaluations/{evaluation_id}/manual/",
        {"reason": "ثبت اشتباه ماه"},
        format="json",
        **scope,
    )
    assert deleted.status_code == 204
    assert not MonthlyEvaluation.objects.filter(pk=evaluation_id).exists()
    soft_deleted = MonthlyEvaluation.all_objects.get(pk=evaluation_id)
    assert soft_deleted.is_deleted

    audit = AuditEvent.objects.get(
        action="evaluation.manual_deleted", entity_id=str(evaluation_id)
    )
    assert audit.school_id == base_data["school1"].id
    assert audit.metadata["reason"] == "[REDACTED]"

    restored = api_client.post(
        "/api/v1/monthly-evaluations/manual/",
        manual_payload(
            base_data,
            metrics=[{"metric_code": "EDU_02", "value": 3}],
            note="ثبت مجدد",
        ),
        format="json",
        **scope,
    )
    assert restored.status_code == 200
    assert restored.data["result"]["restored"] is True
    assert restored.data["evaluation"]["id"] == str(evaluation_id)
    assert MonthlyEvaluation.objects.get(pk=evaluation_id).metric_scores.count() == 2


@pytest.mark.django_db
def test_manual_evaluation_catalog_matches_framework(api_client, base_data):
    api_client.force_authenticate(base_data["manager"])

    response = api_client.get(
        "/api/v1/monthly-evaluations/catalog/",
        HTTP_X_SCHOOL_ID=str(base_data["school1"].id),
    )

    assert response.status_code == 200
    assert response.data["metric_count"] == len(METRIC_CATALOG) == 74
    assert response.data["score_min"] == 0
    assert response.data["score_max"] == 5
    assert {item["code"] for item in response.data["metrics"]} == set(METRIC_CATALOG)
