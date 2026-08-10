from io import BytesIO

import pytest
from django.core.files.uploadedfile import SimpleUploadedFile
from django.core.management import call_command
from openpyxl import Workbook, load_workbook

from hamamooz.apps.academics.calculations import recalculate_class_term
from hamamooz.apps.academics.models import Assessment, Score
from hamamooz.apps.accounts.models import Role, RoleAssignment, User
from hamamooz.apps.imports.models import ImportJob
from hamamooz.apps.imports.services import EXPECTED_HEADERS
from hamamooz.apps.organizations.models import ClassSection
from hamamooz.apps.reports.models import ReportArchive
from hamamooz.apps.reports.services import generate_report
from hamamooz.apps.students.models import Enrollment, Guardian


def xlsx_upload(rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(EXPECTED_HEADERS[ImportJob.ImportType.STUDENTS])
    for row in rows:
        sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    return SimpleUploadedFile(
        "students.xlsx",
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@pytest.mark.django_db
def test_import_template_download_is_authorized_valid_xlsx(
    api_client, base_data, settings, tmp_path
):
    settings.BASE_DIR = tmp_path
    call_command("generate_import_templates", verbosity=0)
    url = "/api/v1/imports/templates/students/"

    unauthenticated = api_client.get(url)
    assert unauthenticated.status_code == 401

    api_client.force_authenticate(base_data["teacher1"])
    authenticated = api_client.get(url)
    assert authenticated.status_code == 200
    authenticated.close()

    api_client.force_authenticate(base_data["manager"])
    missing = api_client.get("/api/v1/imports/templates/not-a-template/")
    assert missing.status_code == 404

    response = api_client.get(url)
    payload = b"".join(response.streaming_content)
    assert response.status_code == 200
    assert response["Content-Type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert response["Content-Disposition"] == 'attachment; filename="students_template.xlsx"'
    assert len(payload) > 100
    assert payload.startswith(b"PK")
    workbook = load_workbook(BytesIO(payload), read_only=True)
    assert workbook.active.max_column > 0


def create_assessment_via_api(api_client, base_data, title="ارزیابی API"):
    api_client.force_authenticate(base_data["teacher1"])
    response = api_client.post(
        "/api/v1/assessments/",
        {
            "course_offering": str(base_data["offering1"].id),
            "assessment_type": str(base_data["final"].id),
            "title": title,
            "assessment_date": "2026-12-20",
            "max_score": "20",
            "weight": "2",
        },
        format="json",
        HTTP_X_SCHOOL_ID=str(base_data["school1"].id),
    )
    assert response.status_code == 201
    return Assessment.objects.get(pk=response.data["id"])


def save_scores_via_api(api_client, base_data, assessment):
    api_client.force_authenticate(base_data["teacher1"])
    response = api_client.post(
        f"/api/v1/assessments/{assessment.id}/scores/bulk/",
        {
            "entries": [
                {
                    "enrollment": str(enrollment.id),
                    "value": value,
                    "status": Score.Status.PRESENT,
                    "note": "ثبت API",
                }
                for enrollment, value in zip(base_data["enrollments"], ["18", "16"], strict=True)
            ]
        },
        format="json",
        HTTP_X_SCHOOL_ID=str(base_data["school1"].id),
    )
    assert response.status_code == 200


@pytest.mark.django_db
def test_complete_assessment_api_workflow_and_locked_correction(api_client, base_data):
    assessment = create_assessment_via_api(api_client, base_data)
    save_scores_via_api(api_client, base_data, assessment)

    api_client.force_authenticate(base_data["teacher1"])
    response = api_client.post(
        f"/api/v1/assessments/{assessment.id}/submit/",
        format="json",
        HTTP_X_SCHOOL_ID=str(base_data["school1"].id),
    )
    assert response.status_code == 200

    api_client.force_authenticate(base_data["deputy"])
    response = api_client.post(
        f"/api/v1/assessments/{assessment.id}/approve/",
        format="json",
        HTTP_X_SCHOOL_ID=str(base_data["school1"].id),
    )
    assert response.status_code == 200
    response = api_client.post(
        f"/api/v1/assessments/{assessment.id}/lock/",
        format="json",
        HTTP_X_SCHOOL_ID=str(base_data["school1"].id),
    )
    assert response.status_code == 200

    score = Score.objects.get(assessment=assessment, enrollment=base_data["enrollments"][0])
    response = api_client.post(
        f"/api/v1/scores/{score.id}/correct-locked/",
        {
            "value": "19",
            "status": Score.Status.PRESENT,
            "note": "اصلاح API",
            "reason": "اصلاح مستند نمره",
        },
        format="json",
        HTTP_X_SCHOOL_ID=str(base_data["school1"].id),
    )
    assert response.status_code == 200
    assert response.data["revision"] == 2

    recalculate_class_term(base_data["class1"], base_data["term"])
    response = api_client.get(
        f"/api/v1/course-offerings/{base_data['offering1'].id}/results/",
        HTTP_X_SCHOOL_ID=str(base_data["school1"].id),
    )
    assert response.status_code == 200
    assert len(response.data) == 2


@pytest.mark.django_db
def test_assessment_reject_update_and_delete_api_paths(api_client, base_data):
    assessment = create_assessment_via_api(api_client, base_data, title="قابل رد")
    save_scores_via_api(api_client, base_data, assessment)
    api_client.force_authenticate(base_data["teacher1"])
    response = api_client.patch(
        f"/api/v1/assessments/{assessment.id}/",
        {"title": "عنوان ویرایش‌شده"},
        format="json",
        HTTP_X_SCHOOL_ID=str(base_data["school1"].id),
    )
    assert response.status_code == 200
    response = api_client.post(
        f"/api/v1/assessments/{assessment.id}/submit/",
        format="json",
        HTTP_X_SCHOOL_ID=str(base_data["school1"].id),
    )
    assert response.status_code == 200
    api_client.force_authenticate(base_data["deputy"])
    response = api_client.post(
        f"/api/v1/assessments/{assessment.id}/reject/",
        {"reason": "نیازمند بازبینی کامل"},
        format="json",
        HTTP_X_SCHOOL_ID=str(base_data["school1"].id),
    )
    assert response.status_code == 200

    draft = create_assessment_via_api(api_client, base_data, title="قابل حذف")
    response = api_client.delete(
        f"/api/v1/assessments/{draft.id}/",
        HTTP_X_SCHOOL_ID=str(base_data["school1"].id),
    )
    assert response.status_code == 204


@pytest.mark.django_db
def test_account_role_password_logout_and_deactivation_api(api_client, base_data):
    login = api_client.post(
        "/api/v1/auth/token/",
        {"username": "manager", "password": "Strong-pass-123"},
        format="json",
    )
    assert login.status_code == 200
    api_client.force_authenticate(base_data["manager"])
    logout = api_client.post(
        "/api/v1/auth/logout/",
        {"refresh": login.data["refresh"]},
        format="json",
    )
    assert logout.status_code == 204

    create_user = api_client.post(
        "/api/v1/users/",
        {
            "username": "new-teacher",
            "email": "new-teacher@example.com",
            "password": "Strong-new-pass-123",
            "first_name": "دبیر",
            "last_name": "جدید",
        },
        format="json",
        HTTP_X_SCHOOL_ID=str(base_data["school1"].id),
    )
    assert create_user.status_code == 201
    user = User.objects.get(username="new-teacher")
    role = api_client.post(
        "/api/v1/role-assignments/",
        {
            "user": str(user.id),
            "organization": str(base_data["organization"].id),
            "school": str(base_data["school1"].id),
            "role": Role.TEACHER,
            "is_active": True,
        },
        format="json",
        HTTP_X_SCHOOL_ID=str(base_data["school1"].id),
    )
    assert role.status_code == 201
    deactivate = api_client.post(
        f"/api/v1/users/{user.id}/deactivate/",
        format="json",
        HTTP_X_SCHOOL_ID=str(base_data["school1"].id),
    )
    assert deactivate.status_code == 200

    api_client.force_authenticate(base_data["teacher1"])
    password = api_client.post(
        f"/api/v1/users/{base_data['teacher1'].id}/change_password/",
        {
            "current_password": "Strong-pass-123",
            "new_password": "Another-strong-pass-456",
        },
        format="json",
    )
    assert password.status_code == 204


@pytest.mark.django_db
def test_import_create_duplicate_and_retry_api_paths(api_client, base_data):
    rows = [["0012345689", "ایمپورت", "API", "2012-03-04", "female"]]
    source = xlsx_upload(rows)
    source_bytes = source.read()
    source.seek(0)
    api_client.force_authenticate(base_data["manager"])
    response = api_client.post(
        "/api/v1/imports/",
        {
            "school": str(base_data["school1"].id),
            "import_type": ImportJob.ImportType.COMPREHENSIVE_SCHOOL,
            "source_file": source,
        },
        format="multipart",
        HTTP_X_SCHOOL_ID=str(base_data["school1"].id),
    )
    assert response.status_code == 201
    job = ImportJob.objects.get(pk=response.data["id"])

    duplicate = api_client.post(
        "/api/v1/imports/",
        {
            "school": str(base_data["school1"].id),
            "import_type": ImportJob.ImportType.COMPREHENSIVE_SCHOOL,
            "source_file": SimpleUploadedFile(
                "students.xlsx",
                source_bytes,
                content_type=("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
            ),
        },
        format="multipart",
        HTTP_X_SCHOOL_ID=str(base_data["school1"].id),
    )
    assert duplicate.status_code == 400

    job.status = ImportJob.Status.FAILED
    job.save(update_fields=["status"])
    retry = api_client.post(
        f"/api/v1/imports/{job.id}/retry/",
        format="json",
        HTTP_X_SCHOOL_ID=str(base_data["school1"].id),
    )
    assert retry.status_code == 200
    assert retry.data["status"] == ImportJob.Status.QUEUED


@pytest.mark.django_db
def test_report_preview_archive_and_download_api(api_client, base_data, settings, tmp_path):
    settings.MEDIA_ROOT = tmp_path
    assessment = create_assessment_via_api(api_client, base_data, title="گزارش API")
    save_scores_via_api(api_client, base_data, assessment)
    assessment.status = Assessment.Status.LOCKED
    assessment.save(update_fields=["status"])
    payload = {
        "term": str(base_data["term"].id),
        "report_type": ReportArchive.ReportType.STUDENT_REPORT_CARD,
        "enrollment": str(base_data["enrollments"][0].id),
    }
    api_client.force_authenticate(base_data["manager"])
    preview = api_client.post(
        "/api/v1/reports/preview/",
        payload,
        format="json",
        HTTP_X_SCHOOL_ID=str(base_data["school1"].id),
    )
    assert preview.status_code == 200
    assert "کارنامه تحصیلی" in preview.data["html"]
    create = api_client.post(
        "/api/v1/reports/",
        payload,
        format="json",
        HTTP_X_SCHOOL_ID=str(base_data["school1"].id),
    )
    assert create.status_code == 201
    report = ReportArchive.objects.get(pk=create.data["id"])
    generate_report(report.id)
    download = api_client.get(
        f"/api/v1/reports/{report.id}/download/",
        HTTP_X_SCHOOL_ID=str(base_data["school1"].id),
    )
    assert download.status_code == 200
    assert download["Content-Type"] == "application/pdf"


@pytest.mark.django_db
def test_guardian_link_and_enrollment_action_api_paths(api_client, base_data):
    api_client.force_authenticate(base_data["manager"])
    guardian_response = api_client.post(
        "/api/v1/guardians/",
        {
            "organization": str(base_data["organization"].id),
            "national_id": "0012345670",
            "first_name": "ولی",
            "last_name": "آزمون",
            "phone_primary": "09120000000",
        },
        format="json",
        HTTP_X_SCHOOL_ID=str(base_data["school1"].id),
    )
    assert guardian_response.status_code == 201
    guardian = Guardian.objects.get(pk=guardian_response.data["id"])
    link = api_client.post(
        f"/api/v1/students/{base_data['students'][0].id}/guardians/",
        {
            "guardian": str(guardian.id),
            "relationship": "father",
            "is_primary": True,
            "can_pick_up": True,
        },
        format="json",
        HTTP_X_SCHOOL_ID=str(base_data["school1"].id),
    )
    assert link.status_code == 201

    new_class = ClassSection.objects.create(
        school=base_data["school1"],
        academic_year=base_data["year"],
        grade_level=base_data["grade"],
        code="7-api",
        title="هفتم API",
        capacity=10,
    )
    enrollment = base_data["enrollments"][0]
    changed = api_client.post(
        f"/api/v1/enrollments/{enrollment.id}/change-class/",
        {"class_section": str(new_class.id), "reason": "تغییر کلاس API"},
        format="json",
        HTTP_X_SCHOOL_ID=str(base_data["school1"].id),
    )
    assert changed.status_code == 200

    status_response = api_client.post(
        f"/api/v1/enrollments/{base_data['enrollments'][1].id}/change-status/",
        {"status": Enrollment.Status.WITHDRAWN, "date": "2026-11-01", "reason": "خروج"},
        format="json",
        HTTP_X_SCHOOL_ID=str(base_data["school1"].id),
    )
    assert status_response.status_code == 200

    RoleAssignment.objects.create(
        user=base_data["manager"],
        organization=base_data["organization"],
        school=base_data["school2"],
        role=Role.SCHOOL_MANAGER,
    )
    transferred = api_client.post(
        f"/api/v1/enrollments/{enrollment.id}/transfer/",
        {
            "school": str(base_data["school2"].id),
            "grade_level": str(base_data["grade"].id),
            "class_section": str(base_data["class2"].id),
            "student_number": "api-transfer",
            "transfer_date": "2026-12-01",
            "reason": "انتقال API",
        },
        format="json",
        HTTP_X_SCHOOL_ID=str(base_data["school1"].id),
    )
    assert transferred.status_code == 201
