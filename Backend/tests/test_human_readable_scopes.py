from io import BytesIO
from types import SimpleNamespace
from urllib.parse import unquote

import pytest
from django.core.files.base import ContentFile
from openpyxl import load_workbook

from hamamooz.apps.accounts.models import Role, RoleAssignment
from hamamooz.apps.attendance.validators import attendance_evidence_upload_to
from hamamooz.apps.imports.models import ImportJob
from hamamooz.apps.organizations.models import School
from hamamooz.apps.reports.models import ReportArchive
from hamamooz.apps.reports.services import render_report_html


@pytest.mark.django_db
def test_scope_resources_expose_human_names_without_replacing_writable_references(
    api_client, base_data
):
    api_client.force_authenticate(base_data["manager"])
    headers = {"HTTP_X_SCHOOL_ID": str(base_data["school1"].id)}

    schools = api_client.get("/api/v1/schools/", **headers)
    students = api_client.get("/api/v1/students/", **headers)
    classes = api_client.get("/api/v1/classes/", **headers)

    assert schools.status_code == students.status_code == classes.status_code == 200
    assert schools.data["results"][0]["organization_name"] == base_data["organization"].name
    assert students.data["results"][0]["organization_name"] == base_data["organization"].name
    assert classes.data["results"][0]["school_name"] == base_data["school1"].name
    assert classes.data["results"][0]["organization_name"] == base_data["organization"].name
    assert classes.data["results"][0]["academic_year_title"] == base_data["year"].title
    assert classes.data["results"][0]["grade_title"] == base_data["grade"].title


@pytest.mark.django_db
def test_dashboard_school_summary_contains_names_but_no_school_identifier(api_client, base_data):
    api_client.force_authenticate(base_data["manager"])
    response = api_client.get(
        "/api/v1/dashboard/summary/",
        HTTP_X_SCHOOL_ID=str(base_data["school1"].id),
    )

    assert response.status_code == 200
    school = response.data["students_by_school"][0]
    assert school["school_name"] == base_data["school1"].name
    assert school["organization_name"] == base_data["organization"].name
    assert "school_id" not in school


@pytest.mark.django_db
def test_attendance_report_identities_use_scope_names_without_school_ids(api_client, base_data):
    api_client.force_authenticate(base_data["manager"])
    headers = {"HTTP_X_SCHOOL_ID": str(base_data["school1"].id)}

    classroom = api_client.get(
        "/api/v1/attendance-reports/class/",
        {
            "class_section": str(base_data["class1"].id),
            "academic_year": str(base_data["year"].id),
        },
        **headers,
    )
    school = api_client.get(
        "/api/v1/attendance-reports/school/",
        {
            "school": str(base_data["school1"].id),
            "academic_year": str(base_data["year"].id),
        },
        **headers,
    )

    assert classroom.status_code == school.status_code == 200
    assert classroom.data["class_section"]["school_name"] == base_data["school1"].name
    assert classroom.data["class_section"]["organization_name"] == base_data["organization"].name
    assert "school" not in classroom.data["class_section"]
    assert school.data["school"]["name"] == base_data["school1"].name
    assert school.data["school"]["organization_name"] == base_data["organization"].name
    assert "id" not in school.data["school"]


@pytest.mark.django_db
def test_duplicate_school_names_are_resolved_by_the_hidden_selected_reference(
    api_client, base_data
):
    duplicate = School.objects.create(
        organization=base_data["organization"],
        code="duplicate-name",
        name=base_data["school1"].name,
    )
    RoleAssignment.objects.create(
        user=base_data["manager"],
        organization=base_data["organization"],
        role=Role.ORGANIZATION_ADMIN,
    )
    api_client.force_authenticate(base_data["manager"])

    response = api_client.get(
        "/api/v1/schools/",
        {"search": base_data["school1"].name},
    )

    assert response.status_code == 200
    selected = next(
        item for item in response.data["results"] if item["id"] == str(base_data["school1"].id)
    )
    assert selected["display_name"] != next(
        item["display_name"] for item in response.data["results"] if item["id"] == str(duplicate.id)
    )
    assert base_data["school1"].name in selected["display_name"]
    assert str(base_data["school1"].id) not in selected["display_name"]


@pytest.mark.django_db
def test_report_archive_and_preview_expose_scope_names(api_client, base_data):
    report = ReportArchive.objects.create(
        organization=base_data["organization"],
        school=base_data["school1"],
        academic_year=base_data["year"],
        term=base_data["term"],
        report_type=ReportArchive.ReportType.STUDENT_REPORT_CARD,
        enrollment=base_data["enrollments"][0],
        requested_by=base_data["manager"],
    )
    api_client.force_authenticate(base_data["manager"])

    response = api_client.get(
        "/api/v1/reports/",
        HTTP_X_SCHOOL_ID=str(base_data["school1"].id),
    )

    assert response.status_code == 200
    item = next(item for item in response.data["results"] if item["id"] == str(report.id))
    assert item["organization_name"] == base_data["organization"].name
    assert item["school_name"] == base_data["school1"].name


def test_report_preview_html_contains_collection_and_school_names_without_scope_ids():
    organization_id = "11111111-1111-4111-8111-111111111111"
    school_id = "22222222-2222-4222-8222-222222222222"
    snapshot = {
        "reports": [
            {
                "organization": {"name": "مجموعه آفتاب"},
                "school": {
                    "name": "مدرسه بهار",
                    "branch": "",
                    "address": "",
                    "phone": "",
                    "manager": "",
                    "logo_url": "",
                },
                "student": {
                    "full_name": "دانش‌آموز نمونه",
                    "national_id": "",
                    "student_number": "",
                    "photo_url": "",
                },
                "academic": {"year": "", "term": "", "grade": "", "class": ""},
                "subjects": [],
                "summary": {
                    "average": None,
                    "class_rank": None,
                    "passed": True,
                    "status_label": "قبول",
                    "formula_version": "v1",
                },
            }
        ]
    }

    html = render_report_html(snapshot, preview=True)

    assert "مجموعه آفتاب" in html
    assert "مدرسه بهار" in html
    assert organization_id not in html
    assert school_id not in html


@pytest.mark.django_db
def test_report_download_filename_uses_names_instead_of_internal_identifier(
    api_client, base_data, settings, tmp_path
):
    settings.MEDIA_ROOT = tmp_path
    report = ReportArchive.objects.create(
        organization=base_data["organization"],
        school=base_data["school1"],
        academic_year=base_data["year"],
        term=base_data["term"],
        report_type=ReportArchive.ReportType.STUDENT_REPORT_CARD,
        enrollment=base_data["enrollments"][0],
        requested_by=base_data["manager"],
        status=ReportArchive.Status.COMPLETED,
    )
    report.output_file.save("internal.pdf", ContentFile(b"%PDF-1.7"), save=True)
    api_client.force_authenticate(base_data["manager"])

    response = api_client.get(
        f"/api/v1/reports/{report.id}/download/",
        HTTP_X_SCHOOL_ID=str(base_data["school1"].id),
    )
    disposition = unquote(response["Content-Disposition"])

    assert response.status_code == 200
    assert base_data["organization"].name in disposition
    assert base_data["school1"].name in disposition
    assert str(report.id) not in disposition


@pytest.mark.django_db
def test_import_error_excel_contains_scope_names_and_not_scope_uuids(api_client, base_data):
    job = ImportJob.objects.create(
        organization=base_data["organization"],
        school=base_data["school1"],
        import_type=ImportJob.ImportType.STUDENTS,
        status=ImportJob.Status.FAILED,
        requested_by=base_data["manager"],
        checksum="a" * 64,
        source_file=ContentFile(b"fixture", name="fixture.xlsx"),
        errors=[{"row": 2, "message": "خطای نمونه"}],
    )
    api_client.force_authenticate(base_data["manager"])

    response = api_client.get(
        f"/api/v1/imports/{job.id}/errors/",
        HTTP_X_SCHOOL_ID=str(base_data["school1"].id),
    )
    payload = b"".join(response.streaming_content)
    workbook = load_workbook(BytesIO(payload), read_only=True, data_only=True)
    metadata = {row[0]: row[1] for row in workbook["scope"].iter_rows(values_only=True)}

    assert response.status_code == 200
    assert metadata["مجموعه"] == base_data["organization"].name
    assert metadata["مدرسه"] == base_data["school1"].name
    assert str(base_data["organization"].id) not in str(metadata)
    assert str(base_data["school1"].id) not in str(metadata)


def test_attendance_evidence_url_does_not_contain_school_identifier():
    school_id = "22222222-2222-4222-8222-222222222222"
    instance = SimpleNamespace(
        attendance_record=SimpleNamespace(
            session=SimpleNamespace(
                school_id=school_id,
                academic_year_id="33333333-3333-4333-8333-333333333333",
            ),
            enrollment=SimpleNamespace(student_id="44444444-4444-4444-8444-444444444444"),
        )
    )

    path = attendance_evidence_upload_to(instance, "evidence.pdf")

    assert path.startswith("attendance/evidence/")
    assert school_id not in path
